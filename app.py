"""
Cegid Retail — Lead Intelligence Platform v3.0
================================================
Merge de app.py (logique enrichissement + scoring McKinsey par produit)
et cegid_lead_intelligence.py (interface + dashboard territoire + pitch IA).

Architecture :
  Page 1 — Enrichissement Dataset     (Sirene + Pappers + OSM)
  Page 2 — Prospection From Scratch   (API Sirene by NAF)
  Page 3 — Scoring & Top 100          (FIT/produit + TIMING + signaux CRM)
  Page 4 — Dashboard Territoire       (carte Plotly par région/pays)
  Page 5 — Générateur de Pitch IA     (Claude API)
  Page 6 — Analyse Business           (McKinsey insights)

Scoring retenu : moteur app.py (FIT par produit Cegid Y2/Orli/Store Excellence)
enrichi des signaux CRM de cegid_lead_intelligence.py (ABX, orphelin, commercial).

Raison : le scoring par produit est directement actionnable pour les équipes
commerciales Cegid (quel produit pitcher, à quel persona). Les signaux CRM
(ABX=20pts, orphelin détecté) augmentent la précision TIMING sans changer
la logique FIT fondamentale.
"""

from __future__ import annotations
import os, re, io, time, json
from pathlib import Path
from datetime import datetime
import streamlit as st
import pandas as pd
import numpy as np
import requests
from bs4 import BeautifulSoup
import anthropic
import plotly.express as px
import plotly.graph_objects as go
import warnings
warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════
# 0. CONFIG
# ══════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Cegid — Lead Intelligence",
    page_icon="data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCAxMDAgMTAwIj4KICA8cmVjdCB3aWR0aD0iMTAwIiBoZWlnaHQ9IjEwMCIgcng9IjE2IiBmaWxsPSIjMDAzMDgyIi8+CiAgPHRleHQgeD0iNTAiIHk9IjcyIiBmb250LWZhbWlseT0iQXJpYWwgQmxhY2ssIHNhbnMtc2VyaWYiIGZvbnQtc2l6ZT0iNjgiIAogICAgICAgIGZvbnQtd2VpZ2h0PSI5MDAiIGZpbGw9IndoaXRlIiB0ZXh0LWFuY2hvcj0ibWlkZGxlIj5DPC90ZXh0Pgo8L3N2Zz4=",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════
# 0-B. STYLES — Interface cegid_lead_intelligence (Inter + metric cards)
#      enrichie des composants log-box / warn-box de app.py
# ══════════════════════════════════════════════════════════════════

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif !important;
}

/* ── Fond général clair ── */
.main, [data-testid="stAppViewContainer"], [data-testid="stMain"],
.block-container {
    background-color: #f8f9fc !important;
}
.stMarkdown, .stMarkdown p, [data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] span {
    color: #0D1F3C !important;
}

/* ── Sidebar gradient Cegid ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0d1f3c 0%, #003082 100%) !important;
}
[data-testid="stSidebar"], [data-testid="stSidebar"] p,
[data-testid="stSidebar"] span, [data-testid="stSidebar"] label,
[data-testid="stSidebar"] li, [data-testid="stSidebar"] a,
[data-testid="stSidebar"] div, [data-testid="stSidebar"] small,
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] li {
    color: white !important;
}

/* ── Hero header ── */
.main-header {
    background: linear-gradient(135deg, #003082, #1a4fa8);
    padding: 1.8rem 2.5rem; border-radius: 16px;
    margin-bottom: 1.5rem; position: relative; overflow: hidden;
}
.main-header::before {
    content: ''; position: absolute; top: -50%; right: -5%;
    width: 280px; height: 280px;
    background: rgba(255,107,53,.12); border-radius: 50%;
}
.main-header h1, .main-header h2, .main-header p,
.main-header span, .main-header div { color: white !important; }
.main-header .accent { color: #FF6B35 !important; }

/* ── Section titles ── */
.section-title {
    font-size: 1.05rem; font-weight: 700; color: #003082 !important;
    border-bottom: 2px solid #FF6B35; padding-bottom: .4rem;
    margin: 1.4rem 0 .9rem; display: inline-block;
}
.section-header {
    font-size: 1.2rem; font-weight: 700; color: #1a1a2e;
    margin: 24px 0 16px; padding-bottom: 8px; border-bottom: 2px solid #e5e7eb;
}

/* ── Metric cards (supporte variantes couleurs) ── */
.metric-card {
    background: white; border-radius: 12px; padding: 18px 22px;
    border-left: 4px solid #003082;
    box-shadow: 0 2px 8px rgba(0,48,130,.08); margin-bottom: 10px;
}
.metric-card.red    { border-left-color: #e63946; }
.metric-card.green  { border-left-color: #2dc653; }
.metric-card.orange { border-left-color: #f4a261; }
.metric-card.purple { border-left-color: #7c3aed; }
.metric-value { font-size: 1.9rem; font-weight: 700; color: #003082; font-family: 'DM Mono', monospace; }
.metric-label { font-size: 0.82rem; color: #6b7280; margin-top: 4px; }

/* ── Grade badges ── */
.grade-A { background:#2dc653; color:white; padding:3px 10px; border-radius:20px; font-weight:700; font-size:.82rem; }
.grade-B { background:#1a6faf; color:white; padding:3px 10px; border-radius:20px; font-weight:700; font-size:.82rem; }
.grade-C { background:#f4a261; color:white; padding:3px 10px; border-radius:20px; font-weight:700; font-size:.82rem; }
.grade-D { background:#e63946; color:white; padding:3px 10px; border-radius:20px; font-weight:700; font-size:.82rem; }
.abx-badge    { background:#7c3aed; color:white; padding:2px 8px; border-radius:12px; font-size:.75rem; font-weight:600; }
.orphan-badge { background:#e63946; color:white; padding:2px 8px; border-radius:12px; font-size:.75rem; font-weight:600; }

/* ── Pitch box ── */
.pitch-box {
    background: white; border-radius: 12px; padding: 24px;
    border: 1px solid #e5e7eb; box-shadow: 0 2px 8px rgba(0,0,0,.06); margin-top: 12px;
}
.pitch-section { margin-bottom: 20px; }
.pitch-title { font-weight:700; color:#003082; margin-bottom:8px; font-size:.88rem; text-transform:uppercase; letter-spacing:.05em; }
.pitch-content { color:#374151; line-height:1.6; font-size:.92rem; }

/* ── Info / Warn / Log boxes ── */
.info-box {
    background:#eff6ff; border:1px solid #bfdbfe;
    border-radius:10px; padding:.9rem 1.1rem;
    margin:.8rem 0; font-size:.88rem; color:#1e40af !important;
}
.warn-box {
    background:#fffbeb; border:1px solid #fde68a;
    border-radius:10px; padding:.9rem 1.1rem;
    margin:.8rem 0; font-size:.88rem; color:#92400e !important;
}
.log-box {
    background:#f8fafc; border:1px solid #e2e8f0;
    border-radius:8px; padding:.7rem 1rem;
    font-family:'DM Mono',monospace; font-size:.74rem;
    color:#475569 !important; max-height:240px; overflow-y:auto; line-height:1.6;
}

/* ── Boutons Cegid ── */
div[data-testid="stButton"] > button {
    background:#003082 !important; color:white !important;
    border:none !important; border-radius:8px !important;
    font-weight:600 !important; transition:all .2s !important;
}
div[data-testid="stButton"] > button:hover {
    background:#FF6B35 !important; color:white !important;
    transform:translateY(-1px);
}
div[data-testid="stButton"] > button p,
div[data-testid="stButton"] > button span { color:white !important; }
.stProgress > div > div { background:#FF6B35 !important; }
div[data-testid="stTabs"] button { font-weight:600; }

/* ── Sidebar nav — numéro + nom, dot actif ── */
[data-testid="stSidebar"] div[data-testid="stButton"] > button {
    background: rgba(255,255,255,0.04) !important;
    border: none !important;
    border-radius: 8px !important;
    color: rgba(255,255,255,0.6) !important;
    font-size: 0.82rem !important;
    font-weight: 400 !important;
    padding: 9px 12px !important;
    margin-bottom: 2px !important;
    text-align: left !important;
    width: 100% !important;
    min-height: 54px !important;
    white-space: pre-line !important;
    line-height: 1.5 !important;
    transition: all .15s !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] div[data-testid="stButton"] > button:hover {
    background: rgba(255,255,255,0.1) !important;
    color: white !important;
}
[data-testid="stSidebar"] div[data-testid="stButton"] > button p,
[data-testid="stSidebar"] div[data-testid="stButton"] > button span {
    color: inherit !important;
    white-space: pre-line !important;
    text-align: left !important;
    font-size: inherit !important;
}


</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════
# 1. CHARGEMENT CLÉS API
# Même logique que cegid_lead_intelligence.py original :
# anthropic.Anthropic() lit ANTHROPIC_API_KEY automatiquement via os.environ
# On s'assure juste qu'elle y est en la copiant depuis st.secrets si besoin
# ══════════════════════════════════════════════════════════════════

def _get_secret(key):
    """Lit une clé : st.secrets en priorité, puis os.environ, puis .env local."""
    try:
        val = st.secrets[key]
        if val: return str(val)
    except Exception:
        pass
    val = os.environ.get(key, "")
    if val: return val
    try:
        env_path = Path(__file__).parent / ".env"
        if env_path.exists():
            for line in env_path.read_text().splitlines():
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, _, v = line.partition("=")
                    if k.strip() == key: return v.strip()
    except Exception:
        pass
    return ""

PAPPERS_KEY = _get_secret("PAPPERS_API_KEY")
GMAPS_KEY   = _get_secret("GOOGLE_MAPS_API_KEY")
SERPAPI_KEY = _get_secret("SERPAPI_KEY")

# Pour anthropic : on injecte dans os.environ pour que anthropic.Anthropic()
# la trouve automatiquement, exactement comme dans le fichier original
_anthr = _get_secret("ANTHROPIC_API_KEY")
if _anthr:
    os.environ["ANTHROPIC_API_KEY"] = _anthr

# ══════════════════════════════════════════════════════════════════
# 2. CONSTANTES MÉTIER
# ══════════════════════════════════════════════════════════════════

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120",
    "Accept-Language": "fr-FR,fr;q=0.9",
}

NAF_RETAIL = {
    "4771Z":"Habillement","4772A":"Chaussures","4772B":"Maroquinerie",
    "4777Z":"Bijouterie/Horlogerie","4775Z":"Parfumerie/Cosmetiques",
    "4711A":"Superettes","4711B":"Supermarches","4711C":"Hypermarches",
    "4711D":"Hard discount","4711F":"Alimentation generale",
    "4719A":"Grands magasins","4719B":"Commerce detail non alimentaire",
    "4721Z":"Fruits et Legumes","4741Z":"Informatique/Telephonie",
    "4751Z":"Textiles","4759A":"Meubles/Decoration",
    "4778A":"Optique","4778C":"Commerce specialise",
    "4779Z":"Occasion","6420Z":"Holding commerce",
}

NAF_QUERIES = {
    "4771Z": ["ZARA","H&M","MANGO","UNIQLO","CACHE CACHE","KIABI","PROMOD","BERSHKA",
              "PULL AND BEAR","PIMKIE","JENNYFER","ETAM","MORGAN","JULES","CELIO",
              "JACK AND JONES","SANDRO","MAJE","IRO","CLAUDIE PIERLOT","BA&SH","ZADIG"],
    "4772A": ["ANDRE","ERAM","BATA","GEOX","CLARKS","TIMBERLAND","CONVERSE","VANS",
              "BIRKENSTOCK","MEPHISTO","FREE LANCE","MINELLI","BOCAGE","JONAK","KICKERS"],
    "4772B": ["LONGCHAMP","LANCEL","LE TANNEUR","LACOSTE","FURLA","COCCINELLE",
              "MICHAEL KORS","COACH","KATE SPADE","SAMSONITE","DELSEY"],
    "4777Z": ["HISTOIRE D OR","PANDORA","SWAROVSKI","MATY","CLEOR",
              "AGATHA","FOSSIL","SWATCH","OMEGA","CARTIER","VAN CLEEF"],
    "4775Z": ["SEPHORA","MARIONNAUD","NOCIBE","YVES ROCHER","SABON","RITUALS",
              "THE BODY SHOP","LUSH","L OCCITANE","PARAPHARMACIE"],
    "4719A": ["GALERIES LAFAYETTE","PRINTEMPS","LE BON MARCHE","BHV","MONOPRIX"],
    "4719B": ["FNAC","CULTURA","FOOT LOCKER","COURIR","SPORT 2000","GO SPORT"],
    "4778A": ["ALAIN AFFLELOU","OPTIC 2000","KRYS","ATOL","GRAND OPTICAL","OPTICAL CENTER"],
    "4759A": ["HABITAT","BUT","CONFORAMA","MAISONS DU MONDE","JARDILAND",
              "CASTORAMA","LEROY MERLIN","BRICOMARCHE","TRUFFAUT","BOTANIC"],
    "4741Z": ["APPLE","SAMSUNG","ORANGE","SFR","BOUYGUES TELECOM","DARTY","FNAC","BOULANGER"],
    "6420Z": ["INDITEX","H AND M","FAST RETAILING","LVMH","KERING","HERMES","RICHEMONT"],
}
NAF_QUERY_DEFAULT = {
    "4771Z":"vetements mode retail","4775Z":"parfumerie beaute cosmetiques",
    "4772A":"chaussures retail","4777Z":"bijouterie horlogerie","4778A":"optique lunettes",
}

TRANCHES_EFF = {
    "00":"0 salarie","01":"1-2","02":"3-5","03":"6-9","11":"10-19","12":"20-49",
    "21":"50-99","22":"100-199","31":"200-249","32":"250-499","41":"500-999",
    "42":"1 000-1 999","51":"2 000-4 999","52":"5 000-9 999","53":"10 000+",
}

REGIONS_FR = [
    "Auvergne-Rhone-Alpes","Bourgogne-Franche-Comte","Bretagne",
    "Centre-Val de Loire","Corse","Grand Est","Hauts-de-France",
    "Ile-de-France","Normandie","Nouvelle-Aquitaine","Occitanie",
    "Pays de la Loire","Provence-Alpes-Cote d'Azur",
]

REGIONS_INSEE = {
    "Auvergne-Rhone-Alpes":"84","Bourgogne-Franche-Comte":"27","Bretagne":"53",
    "Centre-Val de Loire":"24","Corse":"94","Grand Est":"44","Hauts-de-France":"32",
    "Ile-de-France":"11","Normandie":"28","Nouvelle-Aquitaine":"75","Occitanie":"76",
    "Pays de la Loire":"52","Provence-Alpes-Cote d'Azur":"93",
}

REGIONS_COORDS = {
    "Ile-de-France":             (48.8566,  2.3522),
    "Auvergne-Rhone-Alpes":      (45.7485,  4.8467),
    "Nouvelle-Aquitaine":        (44.8378, -0.5792),
    "Occitanie":                 (43.6047,  1.4442),
    "Provence-Alpes-Cote Azur":  (43.2965,  5.3698),
    "Hauts-de-France":           (50.6292,  3.0573),
    "Grand Est":                 (48.5734,  7.7521),
    "Pays de la Loire":          (47.2184, -1.5536),
    "Bretagne":                  (48.1173, -1.6778),
    "Bourgogne-Franche-Comte":   (47.2805,  5.9993),
    "Normandie":                 (49.1829,  0.3707),
    "Centre-Val de Loire":       (47.7516,  1.6751),
    "Corse":                     (42.0396,  9.0129),
}

PAYS_COORDS = {
    "France":      (46.2276,  2.2137),
    "Belgium":     (50.5039,  4.4699),
    "Switzerland": (46.8182,  8.2275),
    "Monaco":      (43.7384,  7.4246),
    "Luxembourg":  (49.6117,  6.1319),
    "UK":          (52.3555, -1.1743),
    "Germany":     (51.1657, 10.4515),
}

DEPT_REGION = {
    "01":"Auvergne-Rhone-Alpes","03":"Auvergne-Rhone-Alpes","07":"Auvergne-Rhone-Alpes",
    "15":"Auvergne-Rhone-Alpes","26":"Auvergne-Rhone-Alpes","38":"Auvergne-Rhone-Alpes",
    "42":"Auvergne-Rhone-Alpes","43":"Auvergne-Rhone-Alpes","63":"Auvergne-Rhone-Alpes",
    "69":"Auvergne-Rhone-Alpes","73":"Auvergne-Rhone-Alpes","74":"Auvergne-Rhone-Alpes",
    "21":"Bourgogne-Franche-Comte","25":"Bourgogne-Franche-Comte","39":"Bourgogne-Franche-Comte",
    "58":"Bourgogne-Franche-Comte","70":"Bourgogne-Franche-Comte","71":"Bourgogne-Franche-Comte",
    "89":"Bourgogne-Franche-Comte","90":"Bourgogne-Franche-Comte",
    "22":"Bretagne","29":"Bretagne","35":"Bretagne","56":"Bretagne",
    "18":"Centre-Val de Loire","28":"Centre-Val de Loire","36":"Centre-Val de Loire",
    "37":"Centre-Val de Loire","41":"Centre-Val de Loire","45":"Centre-Val de Loire",
    "2A":"Corse","2B":"Corse",
    "08":"Grand Est","10":"Grand Est","51":"Grand Est","52":"Grand Est",
    "54":"Grand Est","55":"Grand Est","57":"Grand Est","67":"Grand Est",
    "68":"Grand Est","88":"Grand Est",
    "02":"Hauts-de-France","59":"Hauts-de-France","60":"Hauts-de-France",
    "62":"Hauts-de-France","80":"Hauts-de-France",
    "75":"Ile-de-France","77":"Ile-de-France","78":"Ile-de-France","91":"Ile-de-France",
    "92":"Ile-de-France","93":"Ile-de-France","94":"Ile-de-France","95":"Ile-de-France",
    "14":"Normandie","27":"Normandie","50":"Normandie","61":"Normandie","76":"Normandie",
    "16":"Nouvelle-Aquitaine","17":"Nouvelle-Aquitaine","19":"Nouvelle-Aquitaine",
    "23":"Nouvelle-Aquitaine","24":"Nouvelle-Aquitaine","33":"Nouvelle-Aquitaine",
    "40":"Nouvelle-Aquitaine","47":"Nouvelle-Aquitaine","64":"Nouvelle-Aquitaine",
    "79":"Nouvelle-Aquitaine","86":"Nouvelle-Aquitaine","87":"Nouvelle-Aquitaine",
    "09":"Occitanie","11":"Occitanie","12":"Occitanie","30":"Occitanie","31":"Occitanie",
    "32":"Occitanie","34":"Occitanie","46":"Occitanie","48":"Occitanie","65":"Occitanie",
    "66":"Occitanie","81":"Occitanie","82":"Occitanie",
    "44":"Pays de la Loire","49":"Pays de la Loire","53":"Pays de la Loire",
    "72":"Pays de la Loire","85":"Pays de la Loire",
    "04":"Provence-Alpes-Cote d'Azur","05":"Provence-Alpes-Cote d'Azur",
    "06":"Provence-Alpes-Cote d'Azur","13":"Provence-Alpes-Cote d'Azur",
    "83":"Provence-Alpes-Cote d'Azur","84":"Provence-Alpes-Cote d'Azur",
}

EXCLUSION_KW = [
    "corner","revendeur","galeries lafayette","printemps","bon marche",
    "wholesale","multimarques","pop-up","stand","e-shop uniquement",
    "pure player","marketplace","amazon",
]

# ── Segmentation officielle Cegid ──────────────────────────────
CEGID_SEGMENTS = {
    "S":  {"stores_min":5,   "stores_max":19,  "ca_min_m":5,   "ca_max_m":19,   "label":"S — 5 à 19 magasins"},
    "M":  {"stores_min":20,  "stores_max":99,  "ca_min_m":20,  "ca_max_m":99,   "label":"M — 20 à 99 magasins"},
    "L":  {"stores_min":100, "stores_max":499, "ca_min_m":100, "ca_max_m":499,  "label":"L — 100 à 499 magasins"},
    "XL": {"stores_min":500, "stores_max":9999,"ca_min_m":500, "ca_max_m":99999,"label":"XL — 500+ magasins"},
}

SECTORS_Y2_PRIORITY = [
    "fashion","luxury","ladieswear","menswear","childrenswear","footwear",
    "beauty","cosmetics","sportswear","accessories","lingerie",
]
SECTORS_Y2_SECONDARY = [
    "professional clothes","jewelry","watches","home linens","specialty food",
    "leather goods","gifts","diy","gardening","decoration","toys",
    "telephony","automotive","office supplies","clubs","hotel","museum","culture",
]
SECTEURS_PRIORITAIRES = [
    "Fashion","Luxury","Ladieswear","Menswear","Childrenswear",
    "Footwear","Beauty & Cosmetics","Sportswear","Accessories","Lingerie"
]
SECTEURS_SECONDAIRES = [
    "Jewelry Watches","Home linens","Specialty Foods","Leather Goods",
    "Gifts & Gadgets","Do it yourself and gardening","Decoration and objects",
    "Toys","Art of entertaining","Automotive","Office supplies",
    "Clubs and stadiums","Culture and Entertainment","General Merchandise",
    "Other","Professional clothes"
]
SECTEURS_CIBLES = SECTEURS_PRIORITAIRES + SECTEURS_SECONDAIRES

NAF_CIBLES = [
    "4771Z","4772A","4772B","4773Z","4774Z","4775Z","4776Z","4777Z",
    "4778A","4778B","4778C","4779Z","4711A","4711B","4711C","4711D",
    "4711E","4711F","4719A","4719B","1411Z","1412Z","1413Z","1414Z",
    "1419Z","1420Z","1431Z","1439Z","1512Z","1520Z","2042Z","3213Z","3230Z"
]

KEYWORDS_FASHION = [
    "clothing","fashion","apparel","textile","footwear","shoes",
    "leather","luxury","cosmetic","beauty","sportswear","lingerie",
    "retail sale","habillement","chaussures","maroquinerie","bijouterie",
    "parfum","sport","accessoires","mode","vetement"
]

PERSONAS = {
    "Cegid Retail Y2": {
        "S":  ["Directeur General","DSI","Directeur Retail"],
        "M":  ["Directeur Informatique","DSI","Directeur Retail","CTO"],
        "L":  ["DSI","Directeur Informatique","Directeur Retail","CTO"],
        "XL": ["DSI","CTO","VP Retail","Directeur Informatique"],
    },
    "Cegid Orli": {
        "all": ["DSI","Directeur General","Directeur Logistique","Supply Chain Manager"],
    },
    "Cegid Store Excellence": {
        "all": ["Directeur Retail","Directeur Operations","VP Retail","Directeur Magasins"],
    },
}

GRADE_COLORS = {"A":"#2dc653","B":"#1a6faf","C":"#f4a261","D":"#e63946"}

# ══════════════════════════════════════════════════════════════════
# 3. UTILITAIRES
# ══════════════════════════════════════════════════════════════════

def dept_to_region(cp):
    if not cp: return ""
    return DEPT_REGION.get(str(cp)[:2], "")

def is_empty(val):
    return val is None or str(val).strip() in ("", "nan", "None", "NaN")

def _get_store_col(df):
    if "No. of Stores" in df.columns and pd.to_numeric(df["No. of Stores"], errors="coerce").notna().sum() > 0:
        return "No. of Stores"
    if "Nb Etablissements" in df.columns:
        return "Nb Etablissements"
    return None

def _get_nb_stores(row):
    for col in ["No. of Stores","Nb Etablissements"]:
        v = row.get(col)
        try:
            if not is_empty(v): return int(float(v))
        except: pass
    return 0

def _get_ca_m(row):
    ca_raw = str(row.get("Annual Revenue","") or "")
    if is_empty(ca_raw): return None
    cleaned = re.sub(r"[^\d.,]", "", ca_raw.replace(" ",""))
    try:
        ca_num = float(cleaned.replace(",","."))
        if ca_num > 100_000:
            ca_num = ca_num / 1_000_000
        return ca_num
    except:
        return None

def normalize_name_dedup(name):
    if not name: return ""
    n = str(name).upper().strip()
    for suffix in [" SAS"," SA"," SARL"," SNC"," SASU"," SCM"," SCI"," GIE",
                   " FRANCE"," BOUTIQUES"," STORES"," SHOP"," RETAIL"," GROUP"," GROUPE"]:
        n = n.replace(suffix, "")
    SKIP = {"LES","LE","LA","L","DE","DU","DES","ET","THE","NEW","LTD","INC","STE","CIE","ETS","SOC"}
    words = [w for w in n.strip().split() if w and len(w) >= 2 and w not in SKIP]
    return " ".join(words[:2]).strip()

def get_persona(segment):
    if segment in ["XS","S"]:
        return "DG", "Directeur General"
    elif segment in ["M","L"]:
        return "DSI", "Directeur des Systemes d'Information"
    else:
        return "DSI+Retail", "DSI et Direction Retail"

# ══════════════════════════════════════════════════════════════════
# 4. ENRICHISSEMENT (Sirene + Pappers + OSM)
# ══════════════════════════════════════════════════════════════════

def sirene_search(name, siren=None):
    r = {"siren":None,"naf":None,"naf_label":None,"adresse":None,
         "nb_etab":None,"effectifs":None,"statut":"Non appelee","ok":False}
    try:
        url    = "https://recherche-entreprises.api.gouv.fr/search"
        params = {"q": siren or name, "per_page": 1}
        resp   = requests.get(url, params=params, headers=HEADERS, timeout=8)
        if resp.status_code != 200:
            r["statut"] = "HTTP {} - verifiez connexion".format(resp.status_code); return r
        data = resp.json()
        if not data.get("results"):
            r["statut"] = "Entreprise non trouvee Sirene"; return r
        c = data["results"][0]; s = c.get("siege", {})
        r.update({
            "siren":    c.get("siren"),
            "naf":      s.get("activite_principale",""),
            "naf_label":s.get("libelle_activite_principale",""),
            "adresse":  s.get("adresse",""),
            "nb_etab":  c.get("nombre_etablissements_ouverts", 0),
            "effectifs":TRANCHES_EFF.get(str(c.get("tranche_effectif_salarie") or ""),""),
            "statut":   "OK — {} etab., effectifs: {}".format(
                c.get("nombre_etablissements_ouverts",0),
                TRANCHES_EFF.get(str(c.get("tranche_effectif_salarie") or ""),"inconnu")),
            "ok": True,
        })
    except Exception as e:
        r["statut"] = "Erreur reseau: {}".format(str(e)[:50])
    return r

def pappers_search(siren, key):
    r = {"ca":None,"dirigeant":None,"forme_jur":None,"creation":None,"statut":"Non appelee","ok":False}
    if not key or not key.strip():
        r["statut"] = "Cle absente — ajoutez PAPPERS_API_KEY dans .env"; return r
    if is_empty(siren):
        r["statut"] = "SIREN manquant"; return r
    try:
        siren_clean = re.sub(r"\D","",str(siren))[:9]
        resp = requests.get("https://api.pappers.fr/v2/entreprise",
            params={"api_token":key.strip(),"siren":siren_clean,"comptes":"true","dirigeants":"true"}, timeout=10)
        if resp.status_code == 401: r["statut"] = "Cle Pappers INVALIDE (401)"; return r
        if resp.status_code == 404: r["statut"] = "Entreprise non trouvee Pappers"; return r
        if resp.status_code != 200: r["statut"] = "Pappers HTTP {}".format(resp.status_code); return r
        data    = resp.json()
        comptes = data.get("comptes_sociaux",[])
        ca = None; ca_year = None
        if comptes:
            ca_raw = comptes[0].get("chiffre_affaires")
            ca_year = comptes[0].get("annee")
            if ca_raw: ca = "{:,} EUR".format(int(ca_raw)).replace(",", " ")
        dirs  = data.get("dirigeants",[])
        d_str = None
        if dirs:
            d = dirs[0]
            d_str = "{} {} ({})".format(d.get("prenom",""),d.get("nom",""),d.get("fonction","")).strip()
        r.update({"ca":ca,"dirigeant":d_str,"forme_jur":data.get("forme_juridique"),
                  "creation": int(str(data.get("date_immatriculation_rcs",""))[:4]) if data.get("date_immatriculation_rcs") else None,
                  "statut":"OK — CA {}{} | Dir: {}".format(ca or "N/D"," ({})".format(ca_year) if ca_year else "",d_str or "N/D"),
                  "ok":True})
    except requests.exceptions.Timeout:
        r["statut"] = "Timeout Pappers"
    except Exception as e:
        r["statut"] = "Erreur Pappers: {}".format(str(e)[:60])
    return r

OSM_SHOP_TYPES = {
    "shop": ["clothes","shoes","jewelry","accessories","cosmetics","beauty",
             "optician","sports","toys","books","electronics","department_store",
             "supermarket","convenience","variety_store","gift","stationery",
             "perfumery","bag","watches","fashion","boutique","general"],
    "amenity": ["cafe","restaurant","fast_food","bakery","ice_cream"],
}
OSM_EXCLUDE_TAGS = {
    "internet_sales":"yes","delivery_only":"yes","shop":"online",
    "office":["yes","company","retail"],"building":"warehouse","landuse":"industrial",
}
OSM_NAME_EXCLUDE = [
    "corner","outlet center","galeries lafayette","printemps","le bon marche",
    "bhv","revendeur","wholesale","entrepot","logistique","siege","headquarters","bureau",
]

def _is_physical_owned_store(element, brand_name):
    tags  = element.get("tags",{})
    name  = tags.get("name","").lower()
    brand = tags.get("brand","").lower()
    shop  = tags.get("shop","").lower()
    brand_lower = brand_name.lower().strip()
    brand_word  = brand_lower.split()[0]
    if brand_word not in name and brand_word not in brand:
        return False, "Nom ne correspond pas"
    for excl in OSM_NAME_EXCLUDE:
        if excl in name: return False, "Nom contient '{}'".format(excl)
    for tag_key, tag_val in OSM_EXCLUDE_TAGS.items():
        osm_val = tags.get(tag_key,"")
        if isinstance(tag_val, list):
            if osm_val in tag_val: return False, "Tag exclu"
        else:
            if osm_val == tag_val: return False, "Tag exclu"
    if shop and shop not in ("online","internet"):
        return True, "Magasin physique (shop={})".format(shop)
    if "lat" in element or element.get("type") in ("node","way"):
        return True, "Point géolocalise"
    return False, "Type inconnu"

def _overpass_count(brand_name):
    r = {"nb_stores":None,"nb_exclu":0,"methode":"OpenStreetMap Overpass","statut":"","ok":False,"details":[]}
    brand = brand_name.strip()
    query = """[out:json][timeout:30];
(
  node["brand"="{b}"]["addr:country"="FR"];
  way["brand"="{b}"]["addr:country"="FR"];
  node["name"~"{b}",i]["shop"]["addr:country"="FR"];
  way["name"~"{b}",i]["shop"]["addr:country"="FR"];
  node["operator"="{b}"]["shop"]["addr:country"="FR"];
);
out tags;""".strip().format(b=brand.replace('"','\"'))
    for endpoint in ["https://overpass-api.de/api/interpreter","https://overpass.kumi.systems/api/interpreter"]:
        try:
            resp = requests.post(endpoint, data={"data":query},
                                 headers={"User-Agent":"CegidRetailLeadIntelligence/3.0"}, timeout=25)
            if resp.status_code != 200: continue
            elements = resp.json().get("elements",[])
            if not elements: continue
            kept, exclu = [], []
            for el in elements:
                valid, reason = _is_physical_owned_store(el, brand)
                name_el = el.get("tags",{}).get("name","inconnu")
                if valid: kept.append(name_el)
                else: exclu.append("{} → {}".format(name_el[:25], reason))
            r.update({"nb_stores":len(kept),"nb_exclu":len(exclu),
                      "statut":"OK — {} OSM : {} magasins, {} exclus".format(len(elements),len(kept),len(exclu)),
                      "details":exclu[:5],"ok":True})
            return r
        except: continue
    r["statut"] = "Overpass indisponible"
    return r

def _scrape_store_locator(brand_name):
    r = {"nb_stores":None,"methode":"Scraping store locator","statut":"","ok":False}
    try:
        query = "{} boutiques magasins physiques France nombre".format(brand_name)
        url   = "https://html.duckduckgo.com/html/?q={}".format(requests.utils.quote(query))
        resp  = requests.get(url, headers=HEADERS, timeout=10)
        soup  = BeautifulSoup(resp.text, "html.parser")
        snippets = [el.get_text() for el in soup.find_all("a", class_="result__snippet")][:8]
        full_txt = " ".join(snippets).lower()
        ECOM_SIGNALS = ["site internet","boutique en ligne","e-commerce","livraison a domicile",
                        "revendeurs agrees","points de vente partenaires","corners","galeries lafayette"]
        patterns = [
            (r"(\d+)\s+(?:boutiques?|magasins?|stores?)\s+(?:en\s+propre|physiques?|exclusifs?)", "magasins en propre"),
            (r"(\d+)\s+(?:boutiques?|magasins?|stores?)\s+en\s+france", "magasins en France"),
            (r"reseau\s+(?:de\s+)?(\d+)\s+(?:boutiques?|magasins?)", "reseau de magasins"),
            (r"(\d+)\s+(?:boutiques?|magasins?|stores?)", "mention boutiques"),
        ]
        for pat, label in patterns:
            for m in re.finditer(pat, full_txt):
                nb = int(m.group(1))
                if not (2 <= nb <= 5000): continue
                start = max(0, m.start()-100); end = min(len(full_txt), m.end()+100)
                ctx = full_txt[start:end]
                if any(sig in ctx for sig in ECOM_SIGNALS): continue
                r.update({"nb_stores":nb,"statut":"OK — {} magasins (pattern: '{}')".format(nb,label),"ok":True})
                return r
        r["statut"] = "Store locator : chiffre non trouve"
    except Exception as e:
        r["statut"] = "Store locator erreur: {}".format(str(e)[:50])
    return r

def count_stores(name, gmaps_key=None, serpapi_key=None):
    r = {"nb_stores":None,"methode":None,"statut":"Non appelee","ok":False}
    if not name: r["statut"] = "Nom manquant"; return r
    res = _overpass_count(name)
    if res["ok"] and res["nb_stores"] and res["nb_stores"] > 0: return res
    res2 = _scrape_store_locator(name)
    if res2["ok"] and res2["nb_stores"]: return res2
    if gmaps_key and gmaps_key.strip():
        try:
            resp = requests.get("https://maps.googleapis.com/maps/api/place/textsearch/json",
                params={"query":"{} boutique magasin France".format(name),"language":"fr","region":"fr","key":gmaps_key.strip()}, timeout=10)
            data = resp.json()
            if data.get("status") in ("OK","ZERO_RESULTS"):
                places  = data.get("results",[])
                brand_w = name.lower().split()[0]
                owned   = [p for p in places if brand_w in p.get("name","").lower()
                           and not any(kw in p.get("name","").lower() for kw in EXCLUSION_KW)]
                return {"nb_stores":len(owned),"methode":"Google Places API",
                        "statut":"OK — {} lieux, {} en propre".format(len(places),len(owned)),"ok":True}
        except: pass
    if serpapi_key and serpapi_key.strip():
        try:
            resp   = requests.get("https://serpapi.com/search",
                params={"engine":"google_maps","q":"{} boutique".format(name),"api_key":serpapi_key.strip(),"hl":"fr","gl":"fr"}, timeout=10)
            places = resp.json().get("local_results",[])
            owned  = [p for p in places if name.lower().split()[0] in p.get("title","").lower()]
            return {"nb_stores":len(owned),"methode":"SerpApi","statut":"OK — {} magasins".format(len(owned)),"ok":True}
        except: pass
    r["statut"] = "Toutes sources epuisees sans resultat"
    return r

gmaps_stores = count_stores  # alias

def enrich_row(row):
    row  = row.copy()
    logs = {"Sirene":"Non necessaire","Pappers":"Non necessaire","Magasins":"Non necessaire"}
    name  = str(row.get("Account Name") or row.get("Nom") or "")
    siren = str(row.get("National ID") or row.get("SIREN") or "")
    if siren in ("nan","None",""): siren = ""
    if any(is_empty(row.get(c)) for c in ["National ID","Industry Code","Nb Etablissements"]) and name:
        res = sirene_search(name, siren or None)
        logs["Sirene"] = res["statut"]
        if res["ok"]:
            if is_empty(siren): row["National ID"] = res["siren"]; siren = str(res["siren"] or "")
            if is_empty(row.get("Industry Code")): row["Industry Code"] = res["naf"]; row["Industry Label"] = res["naf_label"]
            if is_empty(row.get("Nb Etablissements")): row["Nb Etablissements"] = res["nb_etab"]
            if is_empty(row.get("Effectifs")): row["Effectifs"] = res["effectifs"]
            if is_empty(row.get("Adresse Siege")): row["Adresse Siege"] = res["adresse"]
    if any(is_empty(row.get(c)) for c in ["Annual Revenue","Dirigeant"]):
        res = pappers_search(siren, PAPPERS_KEY)
        logs["Pappers"] = res["statut"]
        if res["ok"]:
            if is_empty(row.get("Annual Revenue")): row["Annual Revenue"] = res["ca"]
            if is_empty(row.get("Dirigeant")): row["Dirigeant"] = res["dirigeant"]
            if is_empty(row.get("Forme Juridique")): row["Forme Juridique"] = res["forme_jur"]
    if is_empty(row.get("No. of Stores")) and name:
        res = gmaps_stores(name, GMAPS_KEY or None, SERPAPI_KEY or None)
        detail = res["statut"]
        if res.get("nb_exclu",0) > 0: detail += " | Exclus: {}".format(res.get("nb_exclu",0))
        logs["Magasins"] = detail
        if res["ok"] and res["nb_stores"] is not None:
            row["No. of Stores"] = res["nb_stores"]; row["Source Stores"] = res.get("methode","")
    row["Date Enrichissement"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    return row, logs

# ══════════════════════════════════════════════════════════════════
# 5. PROSPECTION API SIRENE
# ══════════════════════════════════════════════════════════════════

def prospect_by_naf(naf_codes, region=None, min_etab=3, max_results=100):
    results, seen   = [], set()
    collect_target  = max_results * 4
    region_code     = REGIONS_INSEE.get(region) if region and region != "Toutes les regions" else None
    url             = "https://recherche-entreprises.api.gouv.fr/search"
    errors, prog    = [], st.empty()

    def fetch_one(query, naf, label=""):
        found = []
        for page in range(1, 4):
            try:
                params = {"q":query,"per_page":25,"page":page}
                if region_code: params["region"] = region_code
                resp = requests.get(url, params=params, headers=HEADERS, timeout=10)
                if resp.status_code == 429: time.sleep(2); resp = requests.get(url, params=params, headers=HEADERS, timeout=10)
                if resp.status_code != 200: break
                data  = resp.json(); items = data.get("results",[])
                if not items: break
                for c in items:
                    siren = c.get("siren","")
                    if not siren or siren in seen: continue
                    if c.get("etat_administratif") == "F": continue
                    nb = c.get("nombre_etablissements_ouverts") or 0
                    if nb < min_etab: continue
                    s = c.get("siege",{}); cp = s.get("code_postal","")
                    seen.add(siren)
                    found.append({
                        "Account Name":    c.get("nom_complet") or c.get("nom_raison_sociale",""),
                        "National ID":     siren,
                        "Industry Code":   s.get("activite_principale","") or naf,
                        "Industry Label":  NAF_RETAIL.get(s.get("activite_principale","") or naf, label),
                        "Billing Country": "France",
                        "Adresse Siege":   s.get("adresse",""),
                        "Code Postal":     cp,
                        "Ville":           s.get("libelle_commune",""),
                        "Region":          s.get("libelle_region","") or dept_to_region(cp) or (region if region and region != "Toutes les regions" else ""),
                        "Nb Etablissements": nb,
                        "No. of Stores":   None,
                        "Annual Revenue":  None,
                        "Effectifs":       TRANCHES_EFF.get(str(c.get("tranche_effectif_salarie") or ""),""),
                        "Dirigeant":       None,
                        "Account Owner":   None,
                        "ABX Target":      None,
                        "Source":          "API Sirene / data.gouv.fr",
                        "Date extraction": datetime.now().strftime("%Y-%m-%d"),
                    })
                total = data.get("total_results",0)
                if page * 25 >= min(total, 75): break
                time.sleep(0.2)
            except Exception as e:
                errors.append("{}: {}".format(query[:20],str(e)[:40])); break
        return found

    for i, naf in enumerate(naf_codes):
        if len(results) >= collect_target: break
        naf_label = NAF_RETAIL.get(naf,"")
        queries   = NAF_QUERIES.get(naf,[])
        if isinstance(queries, list):
            for enseigne in queries:
                if len(results) >= collect_target: break
                prog.caption("NAF {} ({}/{}) — '{}' — {} prospects...".format(naf,i+1,len(naf_codes),enseigne,len(results)))
                results.extend(fetch_one(enseigne, naf, naf_label)); time.sleep(0.2)
        generic = NAF_QUERY_DEFAULT.get(naf,"")
        if generic and len(results) < collect_target:
            results.extend(fetch_one(generic, naf, naf_label))

    prog.empty()
    if errors:
        with st.expander("Erreurs API ({})".format(len(errors)), expanded=False):
            for e in errors[:10]: st.caption(e)
    if not results: return pd.DataFrame()

    df = pd.DataFrame(results)
    df = df.drop_duplicates(subset=["National ID"], keep="first")
    df["_nom_norm"] = df["Account Name"].apply(normalize_name_dedup)
    df = (df.sort_values("Nb Etablissements", ascending=False)
            .drop_duplicates(subset=["_nom_norm"], keep="first")
            .drop(columns=["_nom_norm"]).reset_index(drop=True))
    names   = df["Account Name"].str.upper().str.strip().tolist()
    to_drop = set()
    for i in range(min(len(names), 200)):
        if i in to_drop: continue
        for j in range(i+1, min(i+15, len(names))):
            if j in to_drop: continue
            a, b = names[i], names[j]
            short, long = (a,b) if len(a)<=len(b) else (b,a)
            if len(short) >= 5 and short in long:
                nb_i = df.iloc[i]["Nb Etablissements"] or 0
                nb_j = df.iloc[j]["Nb Etablissements"] or 0
                to_drop.add(j if nb_i >= nb_j else i)
    if to_drop: df = df.drop(index=list(to_drop)).reset_index(drop=True)
    return df.sort_values("Nb Etablissements", ascending=False).reset_index(drop=True).head(max_results)

# ══════════════════════════════════════════════════════════════════
# 6. SCORING — McKinsey FIT/TIMING par Produit Cegid
#              + Signaux CRM (ABX, Orphelin, Commercial assigné)
# ══════════════════════════════════════════════════════════════════

def _assign_segment(nb_stores, ca_m):
    for seg, rules in CEGID_SEGMENTS.items():
        if nb_stores >= rules["stores_min"] and nb_stores <= rules["stores_max"]: return seg
        if ca_m and ca_m >= rules["ca_min_m"] and ca_m <= rules["ca_max_m"]: return seg
    if nb_stores >= 500 or (ca_m and ca_m >= 500): return "XL"
    return None

def _score_product_fit(row, nb_stores, ca_m, naf, lbl, sect, pays, segment):
    """Score FIT dédié à chaque produit Cegid (0-65 pts chacun)."""
    scores = {}

    # ── Cegid Retail Y2 ──────────────────────────────────────────
    fit_y2 = 0
    NAF_Y2_PREFIXES = ["4771","4772","4777","4775","4759","4741","4719","4778","4751","4711","4779","9319","5510","9102","9001"]
    if any(naf.startswith(p) for p in NAF_Y2_PREFIXES): fit_y2 += 25
    elif any(k in lbl or k in sect for k in SECTORS_Y2_PRIORITY): fit_y2 += 20
    elif any(k in lbl or k in sect for k in SECTORS_Y2_SECONDARY): fit_y2 += 10
    if   nb_stores >= 500: fit_y2 += 25
    elif nb_stores >= 100: fit_y2 += 20
    elif nb_stores >= 20:  fit_y2 += 15
    elif nb_stores >= 5:   fit_y2 += 8
    else:                  fit_y2 -= 10
    if any(g.lower() in pays.lower() for g in ["france","belgique","suisse","luxembourg"]): fit_y2 += 15
    elif any(g.lower() in pays.lower() for g in ["uk","united kingdom","allemagne","espagne","italie"]): fit_y2 += 10
    elif any(g.lower() in pays.lower() for g in ["saudi","emirates","china","japan","singapore"]): fit_y2 += 5
    scores["Cegid Retail Y2"] = max(0, fit_y2)

    # ── Cegid Orli ────────────────────────────────────────────────
    fit_orli = 0
    NAF_ORLI_PREFIXES = ["141","142","143","151","152","2042","3213","3230","4641","4642","4645"]
    if any(naf.replace(".","").startswith(p) for p in NAF_ORLI_PREFIXES): fit_orli += 35
    elif any(k in lbl or k in sect for k in ["fabrication","confection","manufacture"]): fit_orli += 20
    if ca_m and ca_m >= 50: fit_orli += 20
    elif ca_m and ca_m >= 10: fit_orli += 15
    elif ca_m and ca_m >= 3: fit_orli += 8
    elif ca_m: fit_orli -= 5
    if any(g.lower() in pays.lower() for g in ["france","belgique","suisse","luxembourg","monaco","andorre"]): fit_orli += 15
    elif any(g.lower() in pays.lower() for g in ["maroc","tunisie"]): fit_orli += 8
    else: fit_orli -= 10
    scores["Cegid Orli"] = max(0, fit_orli)

    # ── Cegid Store Excellence ────────────────────────────────────
    fit_se = 0
    if   nb_stores >= 500: fit_se += 30
    elif nb_stores >= 100: fit_se += 25
    elif nb_stores >= 20:  fit_se += 20
    elif nb_stores >= 10:  fit_se += 5
    else:                  fit_se -= 15
    SE_SECTORS = ["fashion","luxury","ladieswear","footwear","beauty","sportswear","accessories","lingerie","stations services"]
    if any(k in lbl or k in sect for k in SE_SECTORS): fit_se += 20
    elif any(naf.startswith(p) for p in ["4771","4772","4775","4777"]): fit_se += 15
    SE_GEO = ["france","belgique","suisse","luxembourg","uk","united kingdom","ireland","allemagne"]
    if any(g.lower() in pays.lower() for g in SE_GEO): fit_se += 15
    else: fit_se -= 10
    scores["Cegid Store Excellence"] = max(0, fit_se)

    return scores

def compute_score(row):
    """
    Score Total = meilleur FIT produit (0-65) + TIMING enrichi (0-35)
    Signaux CRM intégrés au TIMING :
      - ABX Target = +15 (signal priorité Cegid)
      - Orphelin (pas de commercial assigné) = +10 (quick win)
      - Commercial assigné actif = +5 (suivi en cours)
    """
    det = []
    naf    = str(row.get("Industry Code","") or "").replace(".","")
    lbl    = str(row.get("Industry Label","") or "").lower()
    sect   = str(row.get("Retail Sector","") or "").lower()
    pays   = str(row.get("Billing Country","") or "France").lower()
    nb     = _get_nb_stores(row)
    ca_m   = _get_ca_m(row)
    segment = _assign_segment(nb, ca_m)
    det.append("Segment Cegid: {}".format(segment or "Hors cible"))

    fit_scores    = _score_product_fit(row, nb, ca_m, naf, lbl, sect, pays, segment)
    best_product  = max(fit_scores, key=fit_scores.get)
    best_fit      = fit_scores[best_product]
    det.append("Meilleur FIT: {} ({})".format(best_product, best_fit))

    # ── TIMING (35 pts) ─────────────────────────────────────────
    timing = 0

    # Taille réseau → deal size
    if   nb >= 500: timing += 15; det.append("Réseau XL 500+ → deal majeur +15")
    elif nb >= 100: timing += 12; det.append("Réseau L 100-499 +12")
    elif nb >= 20:  timing += 8;  det.append("Réseau M 20-99 +8")
    elif nb >= 5:   timing += 4;  det.append("Réseau S 5-19 +4")
    else:           det.append("Réseau <5 → hors cible +0")

    # Qualité data lead
    filled = sum(1 for c in ["National ID","Industry Code","Adresse Siege","Effectifs","Annual Revenue","Dirigeant"]
                 if not is_empty(row.get(c)))
    timing += min(6, filled)
    det.append("Qualité lead {}/6 +{}".format(filled, min(6, filled)))

    # Dirigeant identifié / CA connu
    if not is_empty(row.get("Dirigeant")):
        timing += 4; det.append("Dirigeant identifié +4")
    elif not is_empty(row.get("Annual Revenue")):
        timing += 2; det.append("CA connu +2")

    # ── Signaux CRM (de cegid_lead_intelligence) ─────────────────
    owner  = str(row.get("Account Owner","") or "")
    is_abx = str(row.get("ABX Target","") or "").upper() == "YES"
    is_orphan = (owner == "" or owner.lower() in ("salesforce prod","none","nan",""))

    if is_abx:
        timing += 10; det.append("ABX Target Cegid +10 (priorité stratégique)")
    if is_orphan:
        timing += 5; det.append("Compte orphelin +5 (pas de commercial — quick win)")
    elif owner and not is_orphan:
        timing += 2; det.append("Commercial assigné +2")

    total   = min(100, best_fit + timing)
    grade   = "A" if total >= 75 else ("B" if total >= 55 else ("C" if total >= 35 else "D"))
    prio    = {"A":"Priorite 1","B":"Priorite 2","C":"Priorite 3","D":"Hors scope"}[grade]

    personas_dict = PERSONAS.get(best_product,{})
    persona = personas_dict.get(segment, personas_dict.get("all",["Direction Informatique"]))
    persona_str = " / ".join(persona[:3])

    return {
        "Score Total":            total,
        "Score FIT":              best_fit,
        "Score TIMING":           timing,
        "Grade":                  grade,
        "Priorite":               prio,
        "Produit Recommande":     best_product,
        "Segment Cegid":          segment or "Hors cible",
        "Score Y2":               fit_scores["Cegid Retail Y2"],
        "Score Orli":             fit_scores["Cegid Orli"],
        "Score Store Excellence": fit_scores["Cegid Store Excellence"],
        "Persona Cible":          persona_str,
        "is_abx":                 is_abx,
        "is_orphan":              is_orphan,
        "Detail Score":           " | ".join(det),
    }

def score_df(df):
    df = df.copy()
    df["_nom_norm"] = df["Account Name"].apply(normalize_name_dedup)
    nb_col = "Nb Etablissements" if "Nb Etablissements" in df.columns else df.columns[0]
    df[nb_col] = pd.to_numeric(df[nb_col], errors="coerce").fillna(0)
    df = (df.sort_values(nb_col, ascending=False)
            .drop_duplicates(subset=["_nom_norm"], keep="first")
            .drop(columns=["_nom_norm"]).reset_index(drop=True))
    scores = pd.DataFrame(df.apply(compute_score, axis=1).tolist())
    result = pd.concat([df.reset_index(drop=True), scores], axis=1)
    return result.sort_values("Score Total", ascending=False).reset_index(drop=True)

# ── Chargement et scoring intégré (pour import de base CRM) ──────
@st.cache_data
def load_and_score_crm(file_bytes, filename):
    """Charge et score directement une base CRM (avec Status, ABX Target, Account Owner)."""
    if filename.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes), sep=None, engine="python")
    else:
        df = pd.read_excel(io.BytesIO(file_bytes))
    df["No. of Stores"]  = pd.to_numeric(df.get("No. of Stores",  pd.Series()), errors="coerce").clip(upper=10000)
    df["Annual Revenue"] = pd.to_numeric(df.get("Annual Revenue", pd.Series()), errors="coerce")
    if "Status" in df.columns:
        df = df[df["Status"] != "To Be Deleted"].copy()
    return score_df(df)

# ══════════════════════════════════════════════════════════════════
# 7. GÉNÉRATEUR DE PITCH IA (Claude API)
# ══════════════════════════════════════════════════════════════════

def generate_pitch(account_name, sector, segment, country, revenue, stores, product):
    _, persona_label = get_persona(segment)
    prompt = f"""Tu es un expert commercial senior chez Cegid, leader mondial des solutions POS retail cloud.
Tu dois créer un kit de prospection ultra-personnalisé pour ce prospect :

- Entreprise : {account_name}
- Secteur : {sector}
- Segment : {segment} ({stores} magasins, CA ~{revenue}M euros)
- Pays : {country}
- Produit Cegid recommandé : {product}
- Interlocuteur cible : {persona_label}

Cegid : +1 000 retailers clients, 85 000 magasins connectés, présent dans 75 pays.
Clients références : Hermès, Lacoste, Tiffany & Co., L'Occitane, Tommy Hilfiger.

Génère EXACTEMENT ce kit au format JSON strict (sans markdown, sans backticks) :
{{
  "email_j1": {{
    "objet": "...",
    "corps": "..."
  }},
  "inmail_linkedin": "...",
  "pitch_30s": "...",
  "questions_decouverte": ["question 1", "question 2", "question 3"],
  "objection_probable": "...",
  "reponse_objection": "..."
}}

Règles :
- Email : 5-6 lignes max, accroche personnalisée sur le secteur {sector}, une référence client similaire, un CTA clair
- InMail : 3 lignes max, très direct
- Pitch 30s : naturel, pas commercial, centré sur la problématique du {persona_label}
- Questions : ouvertes, stratégiques, pour qualifier le projet {product}
- Ton adapté au {persona_label}
"""
    client  = anthropic.Anthropic()
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1500,
        messages=[{"role":"user","content":prompt}]
    )
    raw = message.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw); raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)

# ══════════════════════════════════════════════════════════════════
# 8. EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════

def to_excel(df):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Leads")
        ws = w.sheets["Leads"]
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        hf = PatternFill(start_color="003082", end_color="003082", fill_type="solid")
        for ci, cell in enumerate(ws[1], 1):
            cell.fill = hf
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(cell.value or ""))+4, 12), 40)
        ws.freeze_panes = "A2"
        alt = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        for ri in range(2, ws.max_row+1, 2):
            for ci in range(1, ws.max_column+1):
                ws.cell(row=ri, column=ci).fill = alt
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════
# 9. SESSION STATE
# ══════════════════════════════════════════════════════════════════

if "df_prospects" not in st.session_state: st.session_state.df_prospects = None
if "df_scored"    not in st.session_state: st.session_state.df_scored    = None

# ══════════════════════════════════════════════════════════════════
# 10. SIDEBAR
# ══════════════════════════════════════════════════════════════════

# ── Pages définition ──────────────────────────────────────────
PAGES = [
    ("Enrichissement Dataset",   "01", "Compléter votre base"),
    ("Prospection From Scratch", "02", "Trouver de nouveaux prospects"),
    ("Scoring & Top 100",        "03", "Prioriser les meilleurs leads"),
    ("Dashboard Territoire",     "04", "Carte & analyse géo"),
    ("Generateur de Pitch IA",   "05", "Générer les kits de vente"),
    ("Analyse Business",         "06", "Insights & recommandations"),
]
PAGE_NAMES = [p[0] for p in PAGES]

if "page_idx" not in st.session_state:
    st.session_state.page_idx = 0

def go_to(idx):
    st.session_state.page_idx = max(0, min(len(PAGES)-1, idx))

with st.sidebar:
    st.markdown("## Cegid Retail")
    st.markdown("**Lead Intelligence Platform v3.0**")
    st.markdown("---")

    # Navigation — numéro + nom sur deux lignes, pill active orange
    for i, (pname, pnum, pdesc) in enumerate(PAGES):
        is_active = (i == st.session_state.page_idx)
        dot = "●" if is_active else "○"
        label = "{} {}  {}\n{}".format(dot, pnum, pdesc, pname)
        if st.button(label, key="nav_{}".format(i), use_container_width=True):
            go_to(i); st.rerun()

    st.markdown("---")
    st.markdown("**Statut clés API**")
    st.markdown("{} Pappers {}".format("✅" if PAPPERS_KEY else "❌","(active)" if PAPPERS_KEY else "(absente)"))
    st.markdown("{} Google Maps {}".format("✅" if GMAPS_KEY else "⚠️","(active)" if GMAPS_KEY else "(optionnel)"))
    st.markdown("{} SerpApi {}".format("✅" if SERPAPI_KEY else "⚠️","(active)" if SERPAPI_KEY else "(optionnel)"))
    st.markdown("{} Anthropic/Claude {}".format("✅" if os.environ.get("ANTHROPIC_API_KEY") else "❌","(active)" if os.environ.get("ANTHROPIC_API_KEY") else "(requise p.5)"))

page = PAGE_NAMES[st.session_state.page_idx]




# ══════════════════════════════════════════════════════════════════
# ── PAGE 1 : ENRICHISSEMENT DATASET ──────────────────────────────
# ══════════════════════════════════════════════════════════════════
try:
    if "Enrichissement Dataset" in page:
        st.markdown("""<div class="main-header">
            <h1>Cegid Retail — <span class="accent">Enrichissement Dataset</span></h1>
            <p>Completion automatique · Sirene · Pappers · OpenStreetMap · Chaque action est tracee</p>
        </div>""", unsafe_allow_html=True)

        if not PAPPERS_KEY:
            msg = ("Cle Pappers absente → CA et dirigeants non recupérés. "
                   "Ajoutez <b>PAPPERS_API_KEY</b> dans Settings → Secrets." if ON_CLOUD else
                   "Cle Pappers absente → ajoutez <code>PAPPERS_API_KEY=votre_cle</code> dans <code>.env</code>.")
            st.markdown('<div class="warn-box">{}</div>'.format(msg), unsafe_allow_html=True)
        if not GMAPS_KEY and not SERPAPI_KEY:
            st.markdown('<div class="warn-box">Aucune cle Maps → le nb de magasins physiques sera estime via OpenStreetMap/scraping (gratuit).</div>', unsafe_allow_html=True)

        uploaded = st.file_uploader("Deposez votre fichier Excel ou CSV", type=["xlsx","xls","csv"])

        if uploaded:
            try:
                df_in = pd.read_csv(uploaded, sep=None, engine="python") if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
            except Exception as e:
                st.error("Erreur chargement : {}".format(e)); st.stop()

            pct_av = (1 - df_in.isna().mean().mean()) * 100
            c1,c2,c3,c4 = st.columns(4)
            with c1: st.markdown('<div class="metric-card"><div class="metric-value">{:,}</div><div class="metric-label">Lignes total</div></div>'.format(len(df_in)), unsafe_allow_html=True)
            with c2: st.markdown('<div class="metric-card green"><div class="metric-value">{:.0f}%</div><div class="metric-label">Completude actuelle</div></div>'.format(pct_av), unsafe_allow_html=True)
            with c3: st.markdown('<div class="metric-card orange"><div class="metric-value">{:,}</div><div class="metric-label">Lignes a enrichir</div></div>'.format(df_in.isna().any(axis=1).sum()), unsafe_allow_html=True)
            with c4: st.markdown('<div class="metric-card red"><div class="metric-value">{}</div><div class="metric-label">Colonnes incompletes</div></div>'.format((df_in.isna().sum()>0).sum()), unsafe_allow_html=True)

            st.markdown('<div class="section-title">Apercu du fichier</div>', unsafe_allow_html=True)
            st.dataframe(df_in.head(8), use_container_width=True, height=230)

            max_rows = st.number_input("Nb max de lignes a enrichir", 1, len(df_in), min(20, len(df_in)))

            if st.button("Lancer l'enrichissement", use_container_width=True):
                df_out = df_in.copy()
                for _col in df_out.columns:
                    try: df_out[_col] = df_out[_col].where(df_out[_col].isna(), df_out[_col].astype(str))
                    except: df_out[_col] = df_out[_col].astype(object)
                to_enrich = df_out[df_in.isna().any(axis=1)].head(max_rows)
                progress  = st.progress(0)
                status    = st.empty()
                log_area  = st.empty()
                all_logs  = []
                n         = len(to_enrich)

                for i, (idx, row) in enumerate(to_enrich.iterrows()):
                    name = str(row.get("Account Name","") or "")[:45]
                    status.markdown("**Enrichissement** `{}` ({}/{})".format(name, i+1, n))
                    progress.progress((i+1)/n)
                    try:
                        enriched, logs = enrich_row(row)
                        for col, val in enriched.items():
                            safe_val = None if (val is None or (isinstance(val,float) and val!=val)) else str(val)
                            try:
                                if col in df_out.columns: df_out.at[idx, col] = safe_val
                            except: pass
                    except Exception as e:
                        logs = {"Sirene":"Erreur: {}".format(str(e)[:40]),"Pappers":"Non appele","Magasins":"Non appele"}
                    all_logs.append("[{:03d}/{}] {:<40} | S: {} | P: {} | M: {}".format(
                        i+1, n, name, logs["Sirene"][:45], logs["Pappers"][:45], logs["Magasins"][:45]))
                    log_area.markdown('<div class="log-box">' + "<br>".join(all_logs[-10:]) + "</div>", unsafe_allow_html=True)
                    time.sleep(0.5)

                status.markdown("**Enrichissement terminé ✅**")
                progress.progress(1.0)
                pct_ap = (1 - df_out.replace("None",None).isnull().mean().mean()) * 100
                c1,c2,c3 = st.columns(3)
                c1.metric("Completude avant","{:.1f}%".format(pct_av))
                c2.metric("Completude apres","{:.1f}%".format(pct_ap),delta="+{:.1f}%".format(pct_ap-pct_av))
                c3.metric("Lignes traitees", n)
                df_display = df_out.replace("None","").replace("nan","")
                st.markdown('<div class="section-title">Resultats enrichis</div>', unsafe_allow_html=True)
                st.dataframe(df_display, use_container_width=True, height=380)
                st.session_state.df_prospects = df_display
                st.download_button("Telecharger le fichier enrichi (.xlsx)",
                    data=to_excel(df_display),
                    file_name="leads_enrichis_{}.xlsx".format(datetime.now().strftime("%Y%m%d_%H%M")),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
        else:
            st.markdown("""<div style="text-align:center;padding:3rem;background:white;border-radius:12px;border:2px dashed #cbd5e1;margin-top:1rem">
                <div style="font-size:3rem">📂</div>
                <h3 style="color:#64748b;font-weight:600">Deposez votre fichier ici</h3>
                <p style="color:#94a3b8">Excel (.xlsx, .xls) ou CSV acceptes</p>
            </div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════
    # ── PAGE 2 : PROSPECTION FROM SCRATCH ────────────────────────
    # ══════════════════════════════════════════════════════════════

    elif "Prospection From Scratch" in page:
        st.markdown("""<div class="main-header">
            <h1>Cegid Retail — <span class="accent">Prospection From Scratch</span></h1>
            <p>Generation de prospects retail France depuis l'API Sirene (gratuit, illimite)</p>
        </div>""", unsafe_allow_html=True)

        st.markdown('<div class="section-title">Criteres de recherche</div>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns([2,2,1])
        with col1:
            naf_sel = st.multiselect("Codes NAF (secteurs retail)", list(NAF_RETAIL.keys()),
                default=["4771Z","4772A","4777Z","4775Z","4778A"],
                format_func=lambda x: "{} — {}".format(x, NAF_RETAIL[x]))
        with col2:
            region_sel = st.selectbox("Region", ["Toutes les regions"] + REGIONS_FR, index=7)
        with col3:
            max_res  = st.number_input("Nb max", 5, 500, 100)
            min_etab = st.number_input("Nb min etab.", 1, 50, 5,
                help="5+ etablissements = enseigne avec réseau de magasins actif")

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("Tester la connexion API", use_container_width=True):
                with st.spinner("Test..."):
                    try:
                        resp = requests.get("https://recherche-entreprises.api.gouv.fr/search",
                            params={"q":"zara","per_page":1}, headers=HEADERS, timeout=5)
                        if resp.status_code == 200: st.success("API Sirene accessible ✅")
                        else: st.error("API Sirene HTTP {}".format(resp.status_code))
                    except Exception as e:
                        st.error("Connexion echouee : {}".format(e))
        with col_btn2:
            run_prosp = st.button("Lancer la prospection", use_container_width=True, type="primary")

        if run_prosp:
            if not naf_sel: st.warning("Selectionnez au moins un code NAF."); st.stop()
            with st.spinner("Interrogation API Sirene..."):
                df_prosp = prospect_by_naf(naf_sel, region=region_sel, min_etab=min_etab, max_results=max_res)
            if df_prosp.empty:
                st.warning("Aucun prospect trouve avec ces criteres. Elargissez la recherche.")
            else:
                st.session_state.df_prospects = df_prosp  # ← sauvegarde immédiate

        # Affiche les résultats depuis session_state (persistant entre les pages)
        if st.session_state.df_prospects is not None:
            df_prosp = st.session_state.df_prospects
            st.success("{:,} prospects en mémoire ✅".format(len(df_prosp)))
            c1,c2,c3,c4 = st.columns(4)
            with c1: st.markdown('<div class="metric-card"><div class="metric-value">{:,}</div><div class="metric-label">Prospects</div></div>'.format(len(df_prosp)), unsafe_allow_html=True)
            with c2: st.markdown('<div class="metric-card green"><div class="metric-value">{}</div><div class="metric-label">Secteurs NAF</div></div>'.format(df_prosp["Industry Code"].nunique()), unsafe_allow_html=True)
            with c3: st.markdown('<div class="metric-card orange"><div class="metric-value">{:.0f}</div><div class="metric-label">Moy. établissements</div></div>'.format(pd.to_numeric(df_prosp["Nb Etablissements"],errors="coerce").mean()), unsafe_allow_html=True)
            _nb = pd.to_numeric(df_prosp["Nb Etablissements"],errors="coerce")
            with c4: st.markdown('<div class="metric-card"><div class="metric-value">{:,}</div><div class="metric-label">Total établissements</div></div>'.format(int(_nb.sum())), unsafe_allow_html=True)
            st.dataframe(df_prosp, use_container_width=True, height=420)
            col_d1, col_d2 = st.columns(2)
            with col_d1:
                st.download_button("Telecharger la liste (.xlsx)", data=to_excel(df_prosp),
                    file_name="prospects_sirene_{}.xlsx".format(datetime.now().strftime("%Y%m%d")),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            with col_d2:
                if st.button("Effacer la liste", key="clear_prosp"):
                    st.session_state.df_prospects = None
                    st.rerun()

    # ══════════════════════════════════════════════════════════════
    # ── PAGE 3 : SCORING & TOP 100 ────────────────────────────────
    # ══════════════════════════════════════════════════════════════

    elif "Scoring & Top 100" in page:
        st.markdown("""<div class="main-header">
            <h1>Cegid Retail — <span class="accent">Scoring & Top 100</span></h1>
            <p>FIT par produit Cegid (Y2/Orli/Store Excellence) · TIMING enrichi signaux CRM · Framework McKinsey</p>
        </div>""", unsafe_allow_html=True)

        with st.expander("Méthodologie de scoring (cliquez pour voir)", expanded=False):
            st.markdown("""
**Score Total = meilleur FIT produit (0–65 pts) + TIMING enrichi (0–35 pts)**

| Dimension | Critère | Max |
|-----------|---------|-----|
| FIT Produit | Adéquation NAF / secteur avec Y2, Orli, Store Excellence | 25 pts |
| FIT Produit | Taille réseau (nb magasins / segment S-M-L-XL) | 25 pts |
| FIT Produit | Géographie (France/DACH/international) | 15 pts |
| TIMING | Taille réseau → proxy deal size / urgence | 15 pts |
| TIMING | Qualité data lead (champs remplis / 6) | 6 pts |
| TIMING | Dirigeant identifié / CA connu | 4 pts |
| **TIMING CRM** | **ABX Target Cegid** | **+10 pts** |
| **TIMING CRM** | **Compte orphelin (pas de commercial assigné)** | **+5 pts** |

**Grades :** A (≥75) Priorité 1 · B (≥55) Priorité 2 · C (≥35) Priorité 3 · D (<35) Hors scope  
**Produit recommandé :** celui avec le meilleur score FIT → pitch adapté automatiquement.
""")

        tab_source = st.radio("Source des données", ["Prospection (Page 2)","Base CRM existante (avec ABX/Owner)"], horizontal=True)
        df_to_score = None

        if "Prospection" in tab_source:
            if st.session_state.df_prospects is not None:
                df_to_score = st.session_state.df_prospects
                st.success("{:,} prospects charges depuis la page Prospection.".format(len(df_to_score)))
            else:
                st.warning("Lancez d'abord une Prospection (Page 2) ou uploadez un fichier.")

        elif "CRM" in tab_source:
            up_crm = st.file_uploader("Uploadez votre base CRM (Prospects & EX Customer)", type=["xlsx","xls","csv"])
            if up_crm:
                with st.spinner("Chargement et pré-traitement CRM..."):
                    df_to_score = load_and_score_crm(up_crm.read(), up_crm.name)
                    st.session_state.df_scored = df_to_score
                st.success("{:,} comptes charges et scorés.".format(len(df_to_score)))
                st.rerun()


        if df_to_score is not None and not df_to_score.empty and "Score Total" not in df_to_score.columns:
            if st.button("Calculer le scoring", use_container_width=True, type="primary"):
                with st.spinner("Calcul du scoring McKinsey..."):
                    df_s = score_df(df_to_score)
                    st.session_state.df_scored = df_s
                st.rerun()

        if st.session_state.df_scored is not None:
            df_s    = st.session_state.df_scored
            top_n   = st.slider("Afficher le Top", 10, min(200,len(df_s)), min(100,len(df_s)), key="top_n_s3")
            df_top  = df_s.head(top_n)
            n_a = (df_s["Grade"]=="A").sum(); n_b = (df_s["Grade"]=="B").sum()
            n_c = (df_s["Grade"]=="C").sum(); n_d = (df_s["Grade"]=="D").sum()
            n_abx    = df_s["is_abx"].sum()   if "is_abx"    in df_s.columns else 0
            n_orphan = df_s["is_orphan"].sum() if "is_orphan" in df_s.columns else 0

            c1,c2,c3,c4,c5,c6 = st.columns(6)
            with c1: st.markdown('<div class="metric-card green"><div class="metric-value">{}</div><div class="metric-label">Grade A — Priorité 1</div></div>'.format(n_a), unsafe_allow_html=True)
            with c2: st.markdown('<div class="metric-card"><div class="metric-value">{}</div><div class="metric-label">Grade B — Priorité 2</div></div>'.format(n_b), unsafe_allow_html=True)
            with c3: st.markdown('<div class="metric-card orange"><div class="metric-value">{}</div><div class="metric-label">Grade C — Priorité 3</div></div>'.format(n_c), unsafe_allow_html=True)
            with c4: st.markdown('<div class="metric-card red"><div class="metric-value">{}</div><div class="metric-label">Hors scope</div></div>'.format(n_d), unsafe_allow_html=True)
            with c5: st.markdown('<div class="metric-card purple"><div class="metric-value">{}</div><div class="metric-label">ABX Target</div></div>'.format(n_abx), unsafe_allow_html=True)
            with c6: st.markdown('<div class="metric-card red"><div class="metric-value">{}</div><div class="metric-label">Orphelins</div></div>'.format(n_orphan), unsafe_allow_html=True)

            # Charts Plotly
            col_l, col_r = st.columns(2)
            with col_l:
                grade_counts = df_s["Grade"].value_counts().reindex(["A","B","C","D"]).fillna(0)
                fig_g = go.Figure(go.Bar(x=grade_counts.index, y=grade_counts.values,
                    marker_color=[GRADE_COLORS[g] for g in grade_counts.index],
                    text=grade_counts.values.astype(int), textposition="outside"))
                fig_g.update_layout(title="Distribution des Grades", height=300,
                    plot_bgcolor="white", paper_bgcolor="white",
                    margin=dict(t=40,b=20,l=20,r=20), showlegend=False)
                st.plotly_chart(fig_g, use_container_width=True)
            with col_r:
                if "Produit Recommande" in df_s.columns:
                    prod_c = df_s["Produit Recommande"].value_counts()
                    fig_p = px.bar(x=prod_c.values, y=prod_c.index, orientation="h",
                        title="Répartition par Produit Cegid Recommandé",
                        color_discrete_sequence=["#003082"])
                    fig_p.update_layout(height=300, plot_bgcolor="white", paper_bgcolor="white",
                        margin=dict(t=40,b=20,l=20,r=20), yaxis_title="", xaxis_title="Nb comptes")
                    st.plotly_chart(fig_p, use_container_width=True)

            # Scatter FIT vs TIMING
            col_l2, col_r2 = st.columns(2)
            with col_l2:
                sample_size = min(500, len(df_s))
                fig_scatter = px.scatter(
                    df_s.head(sample_size), x="Score FIT", y="Score TIMING", color="Grade",
                    color_discrete_map=GRADE_COLORS, title="FIT vs TIMING par Grade",
                    hover_data=["Account Name","Produit Recommande","Segment Cegid"])
                fig_scatter.update_layout(height=320, plot_bgcolor="white", paper_bgcolor="white",
                    margin=dict(t=40,b=20,l=20,r=20))
                st.plotly_chart(fig_scatter, use_container_width=True)
            with col_r2:
                if "Segment Cegid" in df_s.columns:
                    seg_c = df_s["Segment Cegid"].value_counts().reindex(["S","M","L","XL","Hors cible"],fill_value=0)
                    fig_seg = go.Figure(go.Bar(x=seg_c.index, y=seg_c.values,
                        marker_color=["#003082","#1a6faf","#FF6B35","#ff9f1c","#e63946"],
                        text=seg_c.values.astype(int), textposition="outside"))
                    fig_seg.update_layout(title="Répartition par Segment Cegid (S/M/L/XL)",
                        height=320, plot_bgcolor="white", paper_bgcolor="white",
                        margin=dict(t=40,b=20,l=20,r=20), showlegend=False)
                    st.plotly_chart(fig_seg, use_container_width=True)

            st.markdown('<div class="section-title">Top {} — comptes à contacter en priorité</div>'.format(top_n), unsafe_allow_html=True)
            disp_cols = [c for c in ["Account Name","Billing Country","Ville","Industry Label",
                "Nb Etablissements","Effectifs","Annual Revenue","Dirigeant",
                "Score Total","Grade","Priorite","Produit Recommande","Segment Cegid",
                "Persona Cible","Score Y2","Score Orli","Score Store Excellence",
                "Score FIT","Score TIMING","is_abx","is_orphan","Detail Score"]
                if c in df_top.columns]

            def style_grade_col(val):
                colors = {"A":"#2dc653","B":"#1a6faf","C":"#f4a261","D":"#e63946"}
                return "background-color:{}; color:white; font-weight:bold;".format(colors.get(val,"white"))

            try:    styled = df_top[disp_cols].style.map(style_grade_col, subset=["Grade"])
            except: styled = df_top[disp_cols].style.applymap(style_grade_col, subset=["Grade"])
            st.dataframe(styled, use_container_width=True, height=430)

            cd1, cd2 = st.columns(2)
            with cd1:
                st.download_button("Télécharger Top {} (.xlsx)".format(top_n), data=to_excel(df_top),
                    file_name="top{}_cegid_{}.xlsx".format(top_n, datetime.now().strftime("%Y%m%d")),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with cd2:
                st.download_button("Télécharger tous les prospects scorés (.xlsx)", data=to_excel(df_s),
                    file_name="prospects_scores_{}.xlsx".format(datetime.now().strftime("%Y%m%d")),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    # ══════════════════════════════════════════════════════════════
    # ── PAGE 4 : DASHBOARD TERRITOIRE ────────────────────────────
    # ══════════════════════════════════════════════════════════════

    elif "Dashboard Territoire" in page:
        st.markdown("""<div class="main-header">
            <h1>Cegid Retail — <span class="accent">Dashboard Territoire</span></h1>
            <p>Cartographie des prospects · Taille = nb comptes · Couleur = grade dominant</p>
        </div>""", unsafe_allow_html=True)

        df_map_src = st.session_state.df_scored if st.session_state.df_scored is not None else st.session_state.df_prospects
        if df_map_src is None:
            up_map = st.file_uploader("Uploadez un fichier scorer pour la carte", type=["xlsx","xls","csv"])
            if up_map:
                df_map_src = pd.read_csv(up_map, sep=None, engine="python") if up_map.name.endswith(".csv") else pd.read_excel(up_map)
                if "Grade" not in df_map_src.columns:
                    with st.spinner("Scoring en cours..."):
                        df_map_src = score_df(df_map_src)

        if df_map_src is not None:
            grade_col  = "Grade" if "Grade" in df_map_src.columns else "grade"
            score_col  = "Score Total" if "Score Total" in df_map_src.columns else "score_total"
            orphan_col = "is_orphan"
            if grade_col not in df_map_src.columns:
                st.error("Colonne Grade introuvable. Lancez le scoring d'abord.")
                st.stop()
            df_map_src = df_map_src.copy()
            df_map_src["_grade_unified"] = df_map_src[grade_col].astype(str)
            df_map_src["_score_unified"] = pd.to_numeric(df_map_src[score_col] if score_col in df_map_src.columns else 0, errors="coerce").fillna(0)
            df_map_src["_orphan_unified"] = df_map_src[orphan_col].astype(bool) if orphan_col in df_map_src.columns else False

            col_ctrl1, col_ctrl2 = st.columns(2)
            with col_ctrl1:
                map_grade_filter = st.multiselect("Grades à afficher", ["A","B","C","D"], default=["A","B"], key="map_grade_filt")
            with col_ctrl2:
                map_view = st.radio("Vue carte", ["Par region (France)","Par pays"], horizontal=True)

            df_map = df_map_src[df_map_src["_grade_unified"].isin(map_grade_filter)].copy()

            if map_view == "Par region (France)":
                reg_col = next((c for c in ["Region","Billing State/Province (text only)","Billing State/Province"] if c in df_map.columns), None)
                if reg_col:
                    df_france = df_map[df_map["Billing Country"] == "France"].copy() if "Billing Country" in df_map.columns else df_map.copy()
                    df_france["_region"] = df_france[reg_col].fillna("Non renseignee")
                    region_agg = df_france.groupby("_region").agg(
                        nb_comptes     =("Account Name","count"),
                        grade_dominant =("_grade_unified", lambda x: x.value_counts().index[0] if len(x) > 0 else "D"),
                        nb_orphelins   =("_orphan_unified","sum"),
                        score_moyen    =("_score_unified","mean"),
                    ).reset_index()
                    region_agg["lat"] = region_agg["_region"].map(lambda r: REGIONS_COORDS.get(r,(46.5,2.5))[0])
                    region_agg["lon"] = region_agg["_region"].map(lambda r: REGIONS_COORDS.get(r,(46.5,2.5))[1])
                    region_agg["pct_orphelins"] = (region_agg["nb_orphelins"] / region_agg["nb_comptes"] * 100).round(0)
                    # px.scatter_geo : ne nécessite aucun token Mapbox, fonctionne partout
                    fig_map = px.scatter_geo(
                        region_agg, lat="lat", lon="lon",
                        size="nb_comptes", color="grade_dominant",
                        color_discrete_map=GRADE_COLORS,
                        hover_name="_region",
                        hover_data={"nb_comptes":True,"score_moyen":":.1f","pct_orphelins":True,"lat":False,"lon":False},
                        size_max=40,
                        title="Carte Prospects — Par Région")
                    fig_map.update_geos(
                        scope="europe",
                        center={"lat":46.5,"lon":2.5},
                        projection_scale=4,
                        showland=True, landcolor="#f0f0f0",
                        showcoastlines=True, coastlinecolor="#aaaaaa",
                        showcountries=True, countrycolor="#888888",
                        showframe=False,
                        bgcolor="white",
                    )
                else:
                    st.warning("Colonne Region introuvable dans le fichier. Utilisez la vue Par pays.")
                    map_view = "Par pays"

            if map_view == "Par pays":
                pays_col = "Billing Country" if "Billing Country" in df_map.columns else None
                if pays_col:
                    pays_agg = df_map.groupby(pays_col).agg(
                        nb_comptes     =("Account Name","count"),
                        grade_dominant =("_grade_unified", lambda x: x.value_counts().index[0] if len(x)>0 else "D"),
                        nb_orphelins   =("_orphan_unified","sum"),
                        score_moyen    =("_score_unified","mean"),
                    ).reset_index()
                    pays_agg["lat"] = pays_agg[pays_col].map(lambda p: PAYS_COORDS.get(p,(48.0,2.0))[0])
                    pays_agg["lon"] = pays_agg[pays_col].map(lambda p: PAYS_COORDS.get(p,(48.0,2.0))[1])
                    pays_agg["pct_orphelins"] = (pays_agg["nb_orphelins"] / pays_agg["nb_comptes"] * 100).round(0)
                    fig_map = px.scatter_geo(
                        pays_agg, lat="lat", lon="lon",
                        size="nb_comptes", color="grade_dominant",
                        color_discrete_map=GRADE_COLORS,
                        hover_name=pays_col,
                        hover_data={"nb_comptes":True,"score_moyen":":.1f","pct_orphelins":True,"lat":False,"lon":False},
                        size_max=50,
                        title="Carte Prospects — Par Pays")
                    fig_map.update_geos(
                        scope="europe",
                        center={"lat":47.0,"lon":5.0},
                        projection_scale=3,
                        showland=True, landcolor="#f0f0f0",
                        showcoastlines=True, coastlinecolor="#aaaaaa",
                        showcountries=True, countrycolor="#888888",
                        showframe=False,
                        bgcolor="white",
                    )
                else:
                    st.warning("Colonne Billing Country introuvable."); st.stop()

            fig_map.update_layout(height=560, margin=dict(t=40,b=0,l=0,r=0), paper_bgcolor="white")
            st.plotly_chart(fig_map, use_container_width=True)

            # Tableau stats
            st.markdown("---")
            if map_view == "Par region (France)" and reg_col:
                disp_r = region_agg[["_region","nb_comptes","grade_dominant","pct_orphelins","score_moyen"]].copy()
                disp_r.columns = ["Region","Comptes","Grade dominant","% Orphelins","Score moyen"]
                disp_r = disp_r.sort_values("Score moyen", ascending=False)
                def _sg(val):
                    colors = {"A":"#2dc653","B":"#1a6faf","C":"#f4a261","D":"#e63946"}
                    return "background-color:{}; color:white; font-weight:bold;".format(colors.get(val,"white"))
                try:    st.dataframe(disp_r.style.map(_sg, subset=["Grade dominant"]), use_container_width=True, height=350)
                except: st.dataframe(disp_r.style.applymap(_sg, subset=["Grade dominant"]), use_container_width=True, height=350)
        else:
            st.markdown('<div class="info-box">Importez d\'abord une base de prospects (Page 2 ou 3) pour voir la carte.</div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════
    # ── PAGE 5 : GÉNÉRATEUR DE PITCH IA ──────────────────────────
    # ══════════════════════════════════════════════════════════════

    elif "Generateur de Pitch" in page:
        st.markdown("""<div class="main-header">
            <h1>Cegid Retail — <span class="accent">Générateur de Pitch IA</span></h1>
            <p>Kit de prospection personnalisé · Email J1 · InMail · Pitch 30s · Questions découverte · Powered by Claude</p>
        </div>""", unsafe_allow_html=True)

        if not os.environ.get("ANTHROPIC_API_KEY"):
            st.markdown('<div class="warn-box">Clé Anthropic absente → ajoutez <code>ANTHROPIC_API_KEY=votre_cle</code> dans <code>.env</code> ou dans Streamlit Secrets.</div>', unsafe_allow_html=True)

        df_pitch_src = st.session_state.df_scored
        if df_pitch_src is None:
            up_p = st.file_uploader("Uploadez un fichier scorer (.xlsx)", type=["xlsx","xls","csv"])
            if up_p:
                raw = pd.read_csv(up_p, sep=None, engine="python") if up_p.name.endswith(".csv") else pd.read_excel(up_p)
                if "Score Total" not in raw.columns:
                    with st.spinner("Scoring en cours..."): raw = score_df(raw)
                df_pitch_src = raw; st.session_state.df_scored = raw

        if df_pitch_src is not None:
            grade_col_p = "Grade" if "Grade" in df_pitch_src.columns else "grade"
            score_col_p = "Score Total" if "Score Total" in df_pitch_src.columns else "score_total"
            seg_col_p   = "Segment Cegid" if "Segment Cegid" in df_pitch_src.columns else "segment_imputed"
            prod_col_p  = "Produit Recommande" if "Produit Recommande" in df_pitch_src.columns else None
            abx_col_p   = "is_abx" if "is_abx" in df_pitch_src.columns else None
            orphan_col_p = "is_orphan" if "is_orphan" in df_pitch_src.columns else None

            col_sel1, col_sel2 = st.columns([2,1])
            with col_sel1:
                grade_pitch = st.selectbox("Filtrer par grade", ["Tous","A","B","C","D"])
                pool = df_pitch_src if grade_pitch == "Tous" else df_pitch_src[df_pitch_src[grade_col_p] == grade_pitch]
                pool = pool.sort_values(score_col_p, ascending=False)
                account_options = pool["Account Name"].dropna().unique().tolist()
                selected_account = st.selectbox("Sélectionner un compte", account_options)
            with col_sel2:
                st.markdown("<br>", unsafe_allow_html=True)
                generate_btn = st.button("Générer le Kit", use_container_width=True, type="primary")

            if selected_account:
                row_p = pool[pool["Account Name"] == selected_account].iloc[0]
                grade_val    = str(row_p.get(grade_col_p,"?"))
                score_val    = row_p.get(score_col_p, 0)
                seg_val      = str(row_p.get(seg_col_p,"?"))
                sector_val   = str(row_p.get("Retail Sector", row_p.get("Industry Label","Retail")))
                prod_val     = str(row_p.get(prod_col_p,"Cegid Retail Y2")) if prod_col_p else "Cegid Retail Y2"
                pays_val     = str(row_p.get("Billing Country","France"))
                rev_val      = round(float(str(row_p.get("Annual Revenue",0) or 0).replace(" EUR","").replace(",",".").split()[0]) / 1e6 if row_p.get("Annual Revenue") else 0, 1)
                stores_val   = int(row_p.get("No. of Stores") or row_p.get("Nb Etablissements") or 0)
                is_abx_val   = bool(row_p.get(abx_col_p, False)) if abx_col_p else False
                is_orph_val  = bool(row_p.get(orphan_col_p, False)) if orphan_col_p else False
                _, persona_label = get_persona(seg_val)

                col_i1,col_i2,col_i3,col_i4,col_i5 = st.columns(5)
                with col_i1: st.metric("Grade", grade_val)
                with col_i2: st.metric("Score", "{}/100".format(int(score_val)))
                with col_i3: st.metric("Segment", seg_val)
                with col_i4: st.metric("Produit", prod_val.replace("Cegid ",""))
                with col_i5: st.metric("Persona", persona_label.split()[0])

                grade_html  = '<span class="grade-{}">{}</span>'.format(grade_val, grade_val)
                abx_html    = '<span class="abx-badge">ABX</span>' if is_abx_val else ""
                orphan_html = '<span class="orphan-badge">Orphelin</span>' if is_orph_val else ""
                st.markdown("**{}** &nbsp; {} {} {}".format(selected_account, grade_html, abx_html, orphan_html), unsafe_allow_html=True)
                st.markdown("*Interlocuteur recommandé : **{}***".format(persona_label))
                st.markdown("---")

                if generate_btn:
                    if not os.environ.get("ANTHROPIC_API_KEY"):
                        st.error("Clé Anthropic non configurée — impossible de générer le kit.")
                    else:
                        with st.spinner("Génération du kit pour {}...".format(selected_account)):
                            try:
                                pitch = generate_pitch(selected_account, sector_val, seg_val,
                                                       pays_val, rev_val, stores_val, prod_val)
                                st.markdown('<div class="pitch-box">', unsafe_allow_html=True)

                                st.markdown('<div class="pitch-section">', unsafe_allow_html=True)
                                st.markdown('<div class="pitch-title">📧 Email J1 — Objet : {}</div>'.format(pitch["email_j1"]["objet"]), unsafe_allow_html=True)
                                st.markdown('<div class="pitch-content">{}</div>'.format(pitch["email_j1"]["corps"].replace("\n","<br>")), unsafe_allow_html=True)
                                st.markdown('</div>', unsafe_allow_html=True)
                                st.markdown("---")

                                col_p1, col_p2 = st.columns(2)
                                with col_p1:
                                    st.markdown('<div class="pitch-section"><div class="pitch-title">💼 InMail LinkedIn</div>', unsafe_allow_html=True)
                                    st.markdown('<div class="pitch-content">{}</div></div>'.format(pitch["inmail_linkedin"].replace("\n","<br>")), unsafe_allow_html=True)
                                    st.markdown('<div class="pitch-section"><div class="pitch-title">🎤 Pitch 30 secondes</div>', unsafe_allow_html=True)
                                    st.markdown('<div class="pitch-content">{}</div></div>'.format(pitch["pitch_30s"].replace("\n","<br>")), unsafe_allow_html=True)
                                with col_p2:
                                    st.markdown('<div class="pitch-section"><div class="pitch-title">❓ Questions de découverte</div>', unsafe_allow_html=True)
                                    for qi, q in enumerate(pitch["questions_decouverte"], 1):
                                        st.markdown('<div class="pitch-content"><b>{}.</b> {}</div>'.format(qi, q), unsafe_allow_html=True)
                                    st.markdown('</div>', unsafe_allow_html=True)
                                    st.markdown('<div class="pitch-section"><div class="pitch-title">🛡️ Objection probable → Réponse</div>', unsafe_allow_html=True)
                                    st.markdown('<div class="pitch-content"><em>"{}"</em><br><br>→ {}</div></div>'.format(
                                        pitch["objection_probable"], pitch["reponse_objection"]), unsafe_allow_html=True)

                                st.markdown('</div>', unsafe_allow_html=True)

                                pitch_text = "KIT DE PROSPECTION — {}\n{}\nProduit : {} | Persona : {} | Segment : {} | Secteur : {} | Pays : {}\nScore : {}/100 — Grade {}\n\n".format(
                                    selected_account,"="*60, prod_val, persona_label, seg_val, sector_val, pays_val, int(score_val), grade_val)
                                pitch_text += "EMAIL J1\nObjet : {}\n{}\n\nINMAIL LINKEDIN\n{}\n\nPITCH 30 SECONDES\n{}\n\n".format(
                                    pitch["email_j1"]["objet"],pitch["email_j1"]["corps"],pitch["inmail_linkedin"],pitch["pitch_30s"])
                                pitch_text += "QUESTIONS DE DECOUVERTE\n" + "\n".join("{}.  {}".format(i+1,q) for i,q in enumerate(pitch["questions_decouverte"]))
                                pitch_text += "\n\nOBJECTION\n{}\nREPONSE : {}".format(pitch["objection_probable"],pitch["reponse_objection"])

                                st.download_button("Télécharger ce kit (.txt)", pitch_text.encode("utf-8"),
                                    "kit_{}_{}.txt".format(selected_account.replace(" ","_"), datetime.now().strftime("%Y%m%d")),
                                    "text/plain", use_container_width=True)

                            except Exception as e:
                                st.error("Erreur génération : {}".format(e))
                else:
                    st.markdown("""
                    <div class="pitch-box">
                        <div class="pitch-title">Ce qui sera généré pour {}</div>
                        <div class="pitch-content">
                            <b>Email J1</b> — accroche personnalisée secteur {}, référence client similaire, CTA<br><br>
                            <b>InMail LinkedIn</b> — 3 lignes, ultra-direct pour le {}<br><br>
                            <b>Pitch 30 secondes</b> — centré sur la problématique {}<br><br>
                            <b>3 questions de découverte</b> — pour qualifier le projet {}<br><br>
                            <b>Objection probable + réponse</b> — préparer le RDV
                        </div>
                    </div>
                    """.format(selected_account, sector_val, persona_label, persona_label, prod_val), unsafe_allow_html=True)
        else:
            st.markdown('<div class="info-box">Importez une base de prospects depuis la page <strong>Scoring & Top 100</strong> pour accéder au générateur de kit.</div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════
    # ── PAGE 6 : ANALYSE BUSINESS (Mode Consultant McKinsey) ──────
    # ══════════════════════════════════════════════════════════════

    elif "Analyse Business" in page:
        st.markdown("""<div class="main-header">
            <h1>Cegid Retail — <span class="accent">Analyse Business</span></h1>
            <p>Insights stratégiques sur le marché adressable · Mode Consultant McKinsey</p>
        </div>""", unsafe_allow_html=True)

        df_ana = None
        if st.session_state.df_scored is not None:
            df_ana = st.session_state.df_scored
        elif st.session_state.df_prospects is not None:
            df_ana = st.session_state.df_prospects
        else:
            up_ana = st.file_uploader("Uploadez votre fichier de prospects (scoré ou non)", type=["xlsx","xls","csv"])
            if up_ana:
                df_ana = pd.read_csv(up_ana, sep=None, engine="python") if up_ana.name.endswith(".csv") else pd.read_excel(up_ana)
            else:
                st.markdown('<div class="info-box">Lancez d\'abord une <strong>Prospection (Page 2)</strong> puis un <strong>Scoring (Page 3)</strong>, ou uploadez un fichier Excel ici.</div>', unsafe_allow_html=True)

        if df_ana is not None and not df_ana.empty:
            # Normalise les colonnes Grade/Score selon la source
            grade_col_a = "Grade" if "Grade" in df_ana.columns else ("grade" if "grade" in df_ana.columns else None)
            score_col_a = "Score Total" if "Score Total" in df_ana.columns else ("score_total" if "score_total" in df_ana.columns else None)

            st.markdown('<div class="section-title">Marché adressable — vue d\'ensemble</div>', unsafe_allow_html=True)
            _store_col = _get_store_col(df_ana)
            _store_series = pd.to_numeric(df_ana[_store_col], errors="coerce") if _store_col else pd.Series(dtype=float)
            region_col = next((c for c in ["Region","Billing Country"] if c in df_ana.columns), None)
            n_prio = df_ana[grade_col_a].isin(["A","B"]).sum() if grade_col_a else 0
            c1,c2,c3,c4,c5 = st.columns(5)
            with c1: st.markdown('<div class="metric-card"><div class="metric-value">{:,}</div><div class="metric-label">Prospects total</div></div>'.format(len(df_ana)), unsafe_allow_html=True)
            with c2: st.markdown('<div class="metric-card"><div class="metric-value">{:,}</div><div class="metric-label">Total magasins</div></div>'.format(int(_store_series.sum()) if not _store_series.empty else 0), unsafe_allow_html=True)
            with c3: st.markdown('<div class="metric-card green"><div class="metric-value">{:.0f}</div><div class="metric-label">Médiane réseau</div></div>'.format(_store_series.median() if not _store_series.empty else 0), unsafe_allow_html=True)
            with c4: st.markdown('<div class="metric-card"><div class="metric-value">{}</div><div class="metric-label">Régions couvertes</div></div>'.format(df_ana[region_col].nunique() if region_col else "N/A"), unsafe_allow_html=True)
            with c5: st.markdown('<div class="metric-card orange"><div class="metric-value">{}</div><div class="metric-label">Leads prioritaires A+B</div></div>'.format(n_prio), unsafe_allow_html=True)

            # Répartition par secteur
            st.markdown('<div class="section-title">Répartition par secteur retail (NAF)</div>', unsafe_allow_html=True)
            if "Industry Label" in df_ana.columns:
                sect = df_ana.groupby("Industry Label").agg(nb_prospects=("Account Name","count")).sort_values("nb_prospects",ascending=False).head(10).reset_index()
                sect = sect.reset_index(drop=True)
                sect["#"] = sect.index  # numéro pour l'axe X
                cs1, cs2 = st.columns([3,2])
                with cs1:
                    fig_sect = px.bar(sect, x="#", y="nb_prospects",
                        title="Top 10 secteurs NAF", color_discrete_sequence=["#003082"])
                    fig_sect.update_layout(height=320, plot_bgcolor="white", paper_bgcolor="white",
                        margin=dict(t=40,b=20,l=20,r=20), showlegend=False, xaxis_title="", yaxis_title="Nb prospects")
                    fig_sect.update_xaxes(tickmode="array", tickvals=list(sect["#"]), ticktext=[str(i) for i in sect["#"]])
                    st.plotly_chart(fig_sect, use_container_width=True)
                with cs2:
                    disp_sect = sect[["#","Industry Label","nb_prospects"]].rename(columns={"#":"N°","Industry Label":"Secteur","nb_prospects":"Nb"})
                    st.dataframe(disp_sect, use_container_width=True, height=320)

            # Répartition géographique
            st.markdown('<div class="section-title">Répartition géographique</div>', unsafe_allow_html=True)
            if region_col:
                geo = df_ana[df_ana[region_col].notna()].groupby(region_col).agg(nb=("Account Name","count")).sort_values("nb",ascending=False).reset_index()
                cg1, cg2 = st.columns([3,2])
                with cg1:
                    fig_geo = px.bar(geo, x=region_col, y="nb",
                        color_discrete_sequence=["#FF6B35"], title="Prospects par région/pays")
                    fig_geo.update_layout(height=320, plot_bgcolor="white", paper_bgcolor="white",
                        margin=dict(t=40,b=60,l=20,r=20), xaxis_tickangle=-30, showlegend=False, xaxis_title="", yaxis_title="Nb")
                    st.plotly_chart(fig_geo, use_container_width=True)
                with cg2:
                    st.dataframe(geo.rename(columns={region_col:"Region","nb":"Nb Prospects"}), use_container_width=True, height=300)

            # Segmentation taille réseau
            st.markdown('<div class="section-title">Segmentation par taille de réseau</div>', unsafe_allow_html=True)
            if _store_col:
                nb_s = pd.to_numeric(df_ana[_store_col], errors="coerce").dropna()
                segs = pd.cut(nb_s, bins=[0,4,9,19,49,99,float("inf")],
                    labels=["Micro 1-4","Petite enseigne 5-9","Moyenne enseigne 10-19",
                             "Grande enseigne 20-49","Très grande 50-99","Top account 100+"])
                seg_c = segs.value_counts().sort_index().reset_index()
                seg_c.columns = ["Segment","Nb Prospects"]
                seg_c["% du total"] = (seg_c["Nb Prospects"] / seg_c["Nb Prospects"].sum() * 100).round(1).astype(str) + "%"
                ct1, ct2 = st.columns([2,1])
                with ct1:
                    fig_seg2 = px.bar(seg_c, x="Segment", y="Nb Prospects",
                        color_discrete_sequence=["#003082"], title="Nb de prospects par segment")
                    fig_seg2.update_layout(height=300, plot_bgcolor="white", paper_bgcolor="white",
                        margin=dict(t=40,b=80,l=20,r=20), xaxis_tickangle=-20, showlegend=False)
                    st.plotly_chart(fig_seg2, use_container_width=True)
                    st.caption("Base scoring Cegid : cible primaire = 10+ établissements")
                with ct2:
                    st.dataframe(seg_c, use_container_width=True)

            # ── Insights McKinsey ───────────────────────────────────────
            if grade_col_a:
                st.markdown('<div class="section-title">Synthèse scoring & recommandations stratégiques</div>', unsafe_allow_html=True)
                grade_agg_d = {"nb":("Account Name","count")}
                if score_col_a: grade_agg_d["score_moy"] = (score_col_a,"mean")
                if _store_col:
                    df_ana[_store_col] = pd.to_numeric(df_ana[_store_col], errors="coerce")
                    grade_agg_d["etab_moy"] = (_store_col,"mean")
                grade_s = df_ana.groupby(grade_col_a).agg(**grade_agg_d).round(1).reset_index()
                st.dataframe(grade_s, use_container_width=True)

                n_a = (df_ana[grade_col_a]=="A").sum()
                n_b = (df_ana[grade_col_a]=="B").sum()
                n_abx_ana = df_ana["is_abx"].sum() if "is_abx" in df_ana.columns else 0
                n_orph_ana = df_ana["is_orphan"].sum() if "is_orphan" in df_ana.columns else 0

                df_a = df_ana[df_ana[grade_col_a]=="A"]
                top_sect = "N/A"; top_reg = "N/A"
                if "Industry Label" in df_a.columns and len(df_a):
                    vc = df_a["Industry Label"].value_counts()
                    if len(vc): top_sect = vc.index[0]
                if region_col and len(df_a):
                    vc2 = df_a[region_col].value_counts()
                    if len(vc2): top_reg = vc2.index[0]

                score_moy_a = pd.to_numeric(df_a[score_col_a] if score_col_a else pd.Series([0]),errors="coerce").mean() if len(df_a) else 0
                stores_moy_a = pd.to_numeric(df_a[_store_col] if _store_col else pd.Series([0]),errors="coerce").mean() if len(df_a) else 0

                top_product = "N/A"
                if "Produit Recommande" in df_a.columns and len(df_a):
                    top_product = df_a["Produit Recommande"].value_counts().index[0] if len(df_a["Produit Recommande"].dropna()) else "N/A"

                st.markdown("""
<div class="info-box">
<strong>📊 Insights stratégiques — Recommandations McKinsey</strong><br><br>

<b>Marché adressable prioritaire :</b> <b>{n_a}</b> comptes Grade A (Priorité 1) et <b>{n_b}</b> Grade B (Priorité 2),
soit <b>{total}</b> comptes à traiter en priorité immédiate.<br><br>

<b>Quick wins identifiés :</b> <b>{n_abx}</b> comptes ABX Target Cegid (signaux d'intent forts) et
<b>{n_orph}</b> comptes orphelins (pas de commercial assigné → leads libres à capturer).<br><br>

<b>Segmentation Cegid (règles officielles) :</b><br>
&nbsp;&nbsp;— Segment S (5–19 magasins) → <em>Cegid Retail Y2</em>, interlocuteur : Directeur Général<br>
&nbsp;&nbsp;— Segment M (20–99 magasins) → <em>Y2 + Store Excellence</em>, interlocuteur : DSI + Direction Retail<br>
&nbsp;&nbsp;— Segment L/XL (100+ magasins) → <em>Y2 + Store Excellence</em>, interlocuteur : DSI + CTO + VP Retail<br><br>

<b>Profil type Grade A :</b> score moyen <b>{score:.0f}/100</b>, réseau moyen <b>{etab:.0f} établissements</b>.<br>
<b>Produit dominant Grade A :</b> <b>{product}</b> — concentrer les ressources commerciales sur ce produit.<br><br>

<b>Secteur dominant Grade A :</b> {sect} → adapter les success stories et références clients.<br>
<b>Région prioritaire :</b> {reg} → fort potentiel, à couvrir en priorité par l'équipe terrain.<br><br>

<b>Recommandation d'allocation des ressources :</b><br>
Prioriser les comptes M/L/XL (20+ magasins) qui représentent le meilleur ratio effort/ARR pour Cegid.
Les comptes ABX + Grade A constituent la première vague d'outreach (cycle court, probabilité de conversion élevée).
Les comptes orphelins Grade A/B représentent une opportunité de réallocation commerciale immédiate.
</div>""".format(
                    n_a=n_a, n_b=n_b, total=n_a+n_b,
                    n_abx=n_abx_ana, n_orph=n_orph_ana,
                    score=score_moy_a, etab=stores_moy_a,
                    product=top_product, sect=top_sect, reg=top_reg
                ), unsafe_allow_html=True)

            # Export
            st.markdown('<div class="section-title">Export du rapport</div>', unsafe_allow_html=True)
            st.download_button("Télécharger le rapport complet (.xlsx)",
                data=to_excel(df_ana),
                file_name="analyse_business_cegid_{}.xlsx".format(datetime.now().strftime("%Y%m%d")),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)


except Exception as _page_err:
    import traceback
    st.error("Erreur inattendue — rechargez la page (F5). Détail : {}".format(str(_page_err)[:200]))
    with st.expander("Trace complète (debug)", expanded=False):
        st.code(traceback.format_exc())
    st.info("Si le problème persiste, vérifiez les clés API dans le fichier .env.")
